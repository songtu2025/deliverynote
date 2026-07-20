from contextlib import redirect_stderr
from datetime import datetime
from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

try:
    from delivery_note.cli import main, run_batch
    from delivery_note.pipeline import IMPORT_COLUMNS
except ImportError:
    main = None
    run_batch = None
    IMPORT_COLUMNS = None


class CliTests(unittest.TestCase):
    def test_missing_input_returns_error_exit_code(self):
        self.assertIsNotNone(main, "命令行入口尚未实现")
        if main is None:
            return

        stderr = StringIO()
        with redirect_stderr(stderr):
            exit_code = main(
                [
                    "--delivery",
                    "missing-delivery.xlsx",
                    "--purchase",
                    "missing-purchase.xlsx",
                    "--product-info",
                    "missing-products.xlsx",
                    "--supplier-info",
                    "missing-suppliers.xls",
                    "--position-data",
                    "missing-position.xlsx",
                    "--template",
                    "missing-template.xlsx",
                ]
            )

        self.assertEqual(exit_code, 1)
        self.assertIn("文件不存在", stderr.getvalue())


class RunBatchOutputTests(unittest.TestCase):
    def setUp(self):
        self.assertIsNotNone(run_batch, "批处理入口尚未实现")
        self.assertIsNotNone(IMPORT_COLUMNS, "正式导入字段尚未实现")
        if run_batch is None or IMPORT_COLUMNS is None:
            self.skipTest("批处理入口尚未实现")

    @staticmethod
    def create_inputs(
        directory: Path, purchase_quantity: int
    ) -> tuple[Path, Path, Path, Path, Path, Path]:
        delivery_path = directory / "260717-狂飙-SEEKWAY交货单-发货96箱.xlsx"
        delivery_book = Workbook()
        delivery_sheet = delivery_book.active
        delivery_sheet.title = "汇总"
        delivery_sheet.append([])
        delivery_sheet.append([])
        delivery_sheet.append([])
        delivery_sheet.append(["SKU", "US站", "总计"])
        delivery_sheet.append(["SKU-A", 120, 120])
        delivery_book.save(delivery_path)

        product_path = directory / "产品信息.xlsx"
        product_book = Workbook()
        product_sheet = product_book.active
        product_sheet.append(["SKU", "店铺/站点", "品类A", "锁仓MKSU"])
        product_sheet.append(["SKU-A", "SEEKWAY:US", "水鞋", "锁"])
        product_book.save(product_path)

        purchase_path = directory / "采购需求.xlsx"
        purchase_book = Workbook()
        purchase_sheet = purchase_book.active
        purchase_sheet.append(
            ["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"]
        )
        purchase_sheet.append(
            [
                "待交货",
                "KuangBiao",
                "SKU-A",
                "AMAZON:SEEKWAY:US",
                "水鞋-广州仓",
                purchase_quantity,
            ]
        )
        purchase_book.save(purchase_path)

        supplier_path = directory / "供应商资料.xlsx"
        supplier_book = Workbook()
        supplier_sheet = supplier_book.active
        supplier_sheet.append(["供应商编号", "供应商名称", "状态"])
        supplier_sheet.append(["GYS-023", "KuangBiao", "启用"])
        supplier_book.save(supplier_path)

        position_path = directory / "排查表.xlsx"
        position_book = Workbook()
        position_sheet = position_book.active
        position_sheet.title = "MSKU_视图"
        position_sheet.append(
            ["店铺-站点", "积加SKU", "MSKU", "规模定位", "备货定位", "已下单可售天数"]
        )
        position_sheet.append(["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 90.5])
        position_book.save(position_path)

        template_path = directory / "导入模板.xlsx"
        template_book = Workbook()
        template_sheet = template_book.active
        template_sheet["A1"] = "模板提示"
        template_sheet.merge_cells("A1:G1")
        template_sheet.append(IMPORT_COLUMNS)
        template_sheet.append(
            ["示例仓", "示例供应商", "示例SKU", 1, "示例站点", "", ""]
        )
        template_book.save(template_path)
        return (
            delivery_path,
            purchase_path,
            product_path,
            supplier_path,
            position_path,
            template_path,
        )

    def test_batch_outputs_one_workbook_with_import_and_manual_sheets(self):
        with TemporaryDirectory() as directory:
            directory_path = Path(directory)
            delivery, purchase, product, supplier, position, template = (
                self.create_inputs(directory_path, 80)
            )

            output_path, result = run_batch(
                delivery,
                purchase,
                product,
                supplier,
                position,
                template,
                directory_path / "outputs",
                run_time=datetime(2026, 7, 17, 15, 0, 0),
            )

            self.assertTrue(output_path.is_file())
            output_files = list(output_path.parent.glob("*.xlsx"))
            output_book = load_workbook(output_path, data_only=True)
            import_note = output_book["交货导入"]["F3"].value
            pending_sheet = output_book["待处理导入"]
            pending_row = [pending_sheet.cell(3, column).value for column in range(1, 8)]

        self.assertEqual(result.manual_total, 40)
        self.assertTrue(output_path.name.endswith("_交货处理.xlsx"))
        self.assertEqual(len(output_files), 1)
        self.assertEqual(
            output_book.sheetnames, ["交货导入", "待处理导入"]
        )
        self.assertEqual(import_note, "260717-狂飙-01-96箱")
        self.assertEqual(
            pending_row,
            [
                "水鞋-广州仓",
                "GYS-023",
                "SKU-A",
                40,
                "AMAZON:SEEKWAY:US",
                "260717-狂飙-01-96箱",
                "超出采购未交量：40",
            ],
        )
        self.assertEqual(
            [pending_sheet.cell(3, column).value for column in range(8, 11)],
            ["短尾", "备货", 90.5],
        )

    def test_no_manual_rows_keep_empty_pending_import_sheet(self):
        with TemporaryDirectory() as directory:
            directory_path = Path(directory)
            delivery, purchase, product, supplier, position, template = (
                self.create_inputs(directory_path, 150)
            )

            output_path, result = run_batch(
                delivery,
                purchase,
                product,
                supplier,
                position,
                template,
                directory_path / "outputs",
                run_time=datetime(2026, 7, 17, 15, 0, 0),
            )
            output_files = list(output_path.parent.glob("*.xlsx"))
            output_book = load_workbook(output_path, data_only=True)

        self.assertTrue(output_path.name.endswith("_交货处理.xlsx"))
        self.assertEqual(result.manual_total, 0)
        self.assertEqual(len(output_files), 1)
        self.assertEqual(
            output_book.sheetnames, ["交货导入", "待处理导入"]
        )
        self.assertIsNone(output_book["待处理导入"]["A3"].value)


if __name__ == "__main__":
    unittest.main()
