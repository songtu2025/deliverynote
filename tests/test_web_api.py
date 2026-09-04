import asyncio
from concurrent.futures import ThreadPoolExecutor
from contextlib import suppress
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from threading import Barrier, BrokenBarrierError, Lock
import unittest
from unittest.mock import patch

from httpx2 import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
import pandas as pd
from sqlalchemy import event, select

from delivery_note.pipeline import IMPORT_COLUMNS
from delivery_note.self_operated_inbound import INBOUND_TEMPLATE_COLUMNS
from delivery_note.web.auth import hash_token
from tests.asgi_client import SyncASGIClient

try:
    import delivery_note.web.api as web_api_module
    from delivery_note.web.api import create_app
    from delivery_note.web.models import (
        AuthSession,
        Batch,
        BatchFile,
        ExceptionRecord,
        InputVersion,
        Job,
        PurchaseSyncJob,
        SelfOperatedBatch,
        SelfOperatedInboundSyncJob,
        SplitRecord,
        User,
    )
except ImportError:
    web_api_module = None
    create_app = None
    AuthSession = None
    Batch = None
    BatchFile = None
    ExceptionRecord = None
    InputVersion = None
    PurchaseSyncJob = None
    SelfOperatedBatch = None
    SelfOperatedInboundSyncJob = None
    SplitRecord = None
    User = None


INPUT_KINDS = ("purchase", "product", "supplier", "position", "template")


class WebApiTests(unittest.TestCase):
    def setUp(self):
        self.assertIsNotNone(create_app, "FastAPI 应用尚未实现")
        if create_app is None:
            self.skipTest("FastAPI 应用尚未实现")
        self.directory = TemporaryDirectory()
        root = Path(self.directory.name)
        self.root = root
        self.app = create_app(
            database_url=f"sqlite+pysqlite:///{root / 'test.db'}",
            storage_root=root / "storage",
            bootstrap_admin=("admin", "admin-pass"),
            session_cookie_secure=False,
        )
        self.client = SyncASGIClient(self.app)

    def tearDown(self):
        if hasattr(self, "client"):
            self.client.close()
        if hasattr(self, "app"):
            self.app.state.database.dispose()
        if hasattr(self, "directory"):
            self.directory.cleanup()

    def login(self, username: str, password: str) -> dict[str, str]:
        response = self.client.post(
            "/api/auth/login",
            json={"username": username, "password": password},
        )
        self.assertEqual(response.status_code, 200, response.text)
        return {"Authorization": f"Bearer {response.json()['token']}"}

    def create_operator(self, admin_headers: dict[str, str]) -> dict:
        response = self.client.post(
            "/api/users",
            headers=admin_headers,
            json={
                "username": "operator",
                "password": "operator-pass",
                "role": "operator",
            },
        )
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()

    @staticmethod
    def workbook_bytes(kind: str) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        if kind == "purchase":
            sheet.append(["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"])
            sheet.append(
                [
                    "待交货",
                    "KuangBiao",
                    "SKU-A",
                    "AMAZON:SEEKWAY:US",
                    "水鞋-广州仓",
                    100,
                ]
            )
        elif kind == "product":
            sheet.append(["SKU", "店铺/站点", "品类A", "锁仓MKSU"])
            sheet.append(["SKU-A", "SEEKWAY:US", "水鞋", "锁"])
        elif kind == "supplier":
            sheet.append(["供应商编号", "供应商名称", "状态"])
            sheet.append(["GYS-023", "KuangBiao", "启用"])
        elif kind == "position":
            sheet.title = "MSKU_视图"
            sheet.append(
                [
                    "店铺-站点",
                    "积加SKU",
                    "MSKU",
                    "规模定位",
                    "备货定位",
                    "已下单可售天数",
                ]
            )
            sheet.append(["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 90])
        elif kind == "template":
            sheet["A1"] = "模板提示"
            sheet.merge_cells("A1:G1")
            sheet.append(IMPORT_COLUMNS)
            sheet.append(["示例仓", "示例供应商", "示例SKU", 1, "示例站点", "", ""])
        elif kind == "inbound_template":
            sheet.title = "批量入库"
            sheet.append(INBOUND_TEMPLATE_COLUMNS)
            sheet.append(["示例"] + [""] * (len(INBOUND_TEMPLATE_COLUMNS) - 1))
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def delivery_bytes(quantity: int = 40) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "汇总"
        sheet.append([])
        sheet.append([])
        sheet.append([])
        sheet.append(["SKU", "US站", "总计"])
        sheet.append(["SKU-A", quantity, quantity])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def self_operated_delivery_bytes(quantity: int = 15) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "明细"
        sheet.append([])
        sheet.append([])
        sheet.append([])
        sheet.append(["积加SKU", "实收数量", "站点", "交货单号"])
        sheet.append(["SKU-A", quantity, "US站", "LN2608179025"])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def self_operated_inbound_bytes() -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(INBOUND_TEMPLATE_COLUMNS + ["供应商"])
        values = {column: "" for column in INBOUND_TEMPLATE_COLUMNS}
        values.update(
            {
                "入库单号": "IN-1",
                "入库仓": "自营仓",
                "SKU": "SKU-A",
                "平台站点": "AMAZON:SEEKWAY:US",
                "关联采购单": "PO-20260801",
                "关联交货单/调拨单": "LN2608179025",
                "应收货": 10,
            }
        )
        sheet.append(
            [values[column] for column in INBOUND_TEMPLATE_COLUMNS] + ["KuangBiao"]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def upload_active_versions(self, headers: dict[str, str]) -> dict[str, int]:
        version_ids = {}
        for kind in INPUT_KINDS:
            response = self.client.post(
                f"/api/input-versions/{kind}",
                headers=headers,
                data={"name": f"{kind}-v1", "activate": "true"},
                files={"file": (f"{kind}.xlsx", BytesIO(self.workbook_bytes(kind)))},
            )
            self.assertEqual(response.status_code, 201, response.text)
            version_ids[kind] = response.json()["id"]
        return version_ids

    def get_with_query_count(
        self,
        path: str,
        headers: dict[str, str],
    ):
        statement_count = 0

        def count_statement(*_args):
            nonlocal statement_count
            statement_count += 1

        engine = self.app.state.database.engine
        event.listen(engine, "before_cursor_execute", count_statement)
        try:
            response = self.client.get(path, headers=headers)
        finally:
            event.remove(engine, "before_cursor_execute", count_statement)
        return response, statement_count

    def test_login_and_admin_role_are_enforced(self):
        bad_login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "wrong"},
        )
        self.assertEqual(bad_login.status_code, 401)

        admin_headers = self.login("admin", "admin-pass")
        self.create_operator(admin_headers)
        me = self.client.get("/api/auth/me", headers=admin_headers)
        self.assertEqual(me.status_code, 200)
        self.assertEqual(me.json()["role"], "admin")

        operator_headers = self.login("operator", "operator-pass")
        forbidden = self.client.post(
            "/api/users",
            headers=operator_headers,
            json={
                "username": "another",
                "password": "another-pass",
                "role": "operator",
            },
        )
        self.assertEqual(forbidden.status_code, 403)

        logout = self.client.post("/api/auth/logout", headers=operator_headers)
        self.assertEqual(logout.status_code, 204)
        self.assertEqual(
            self.client.get("/api/auth/me", headers=operator_headers).status_code,
            401,
        )

    def test_login_sets_http_only_strict_cookie_and_rejects_bad_credentials(self):
        bad_login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "wrong"},
        )
        self.assertEqual(bad_login.status_code, 401)
        self.assertNotIn("set-cookie", bad_login.headers)

        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        self.assertEqual(login.status_code, 200, login.text)
        cookie = login.headers["set-cookie"]
        self.assertIn("delivery_note_session=", cookie)
        self.assertIn("HttpOnly", cookie)
        self.assertIn("SameSite=strict", cookie)
        self.assertIn("Path=/", cookie)
        self.assertIn("expires=", cookie.lower())
        self.assertNotIn("; Secure", cookie)
        cookie_expiry = parsedate_to_datetime(
            next(
                attribute.split("=", 1)[1]
                for attribute in cookie.split("; ")
                if attribute.lower().startswith("expires=")
            )
        )
        response_expiry = datetime.fromisoformat(
            login.json()["expires_at"].replace("Z", "+00:00")
        )
        self.assertLessEqual(
            abs(
                (
                    cookie_expiry.astimezone(timezone.utc)
                    - response_expiry.astimezone(timezone.utc)
                ).total_seconds()
            ),
            1,
        )

    def test_session_cookie_secure_environment_is_parsed_strictly(self):
        for configured, expected_secure in (("false", False), ("true", True)):
            with self.subTest(configured=configured), TemporaryDirectory() as directory:
                root = Path(directory)
                with patch.dict(
                    web_api_module.os.environ,
                    {"SESSION_COOKIE_SECURE": configured},
                ):
                    app = create_app(
                        database_url=f"sqlite+pysqlite:///{root / 'cookie.db'}",
                        storage_root=root / "storage",
                        bootstrap_admin=("admin", "admin-pass"),
                    )
                client = SyncASGIClient(app)
                try:
                    response = client.post(
                        "/api/auth/login",
                        json={"username": "admin", "password": "admin-pass"},
                    )
                    self.assertEqual(response.status_code, 200, response.text)
                    self.assertEqual(
                        "; Secure" in response.headers["set-cookie"],
                        expected_secure,
                    )
                finally:
                    client.close()
                    app.state.database.dispose()

        with TemporaryDirectory() as directory:
            root = Path(directory)
            with (
                patch.dict(
                    web_api_module.os.environ,
                    {"SESSION_COOKIE_SECURE": "sometimes"},
                ),
                self.assertRaisesRegex(ValueError, "SESSION_COOKIE_SECURE"),
            ):
                create_app(
                    database_url=f"sqlite+pysqlite:///{root / 'invalid-cookie.db'}",
                    storage_root=root / "storage",
                    auto_migrate_schema=False,
                )

    def test_cookie_auth_supports_me_and_business_requests_with_bearer_priority(self):
        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        self.assertEqual(login.status_code, 200, login.text)
        token = login.json()["token"]
        cookie_headers = {"Cookie": f"delivery_note_session={token}"}

        me = self.client.get("/api/auth/me", headers=cookie_headers)
        self.assertEqual(me.status_code, 200, me.text)
        self.assertEqual(me.json()["username"], "admin")
        batches = self.client.get("/api/batches", headers=cookie_headers)
        self.assertEqual(batches.status_code, 200, batches.text)

        bearer = self.client.get(
            "/api/auth/me",
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(bearer.status_code, 200, bearer.text)
        rejected = self.client.get(
            "/api/auth/me",
            headers={
                "Authorization": "Bearer invalid-bearer",
                "Cookie": f"delivery_note_session={token}",
            },
        )
        self.assertEqual(rejected.status_code, 401, rejected.text)
        self.assertNotIn("set-cookie", rejected.headers)

    def test_cookie_logout_revokes_session_and_clears_cookie(self):
        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        token = login.json()["token"]
        cookie_headers = {"Cookie": f"delivery_note_session={token}"}

        logout = self.client.post("/api/auth/logout", headers=cookie_headers)
        self.assertEqual(logout.status_code, 204, logout.text)
        expired_cookie = logout.headers["set-cookie"]
        self.assertIn("delivery_note_session=", expired_cookie)
        self.assertIn("Max-Age=0", expired_cookie)
        self.assertIn("HttpOnly", expired_cookie)
        self.assertIn("SameSite=strict", expired_cookie)
        with self.app.state.database.session() as session:
            auth_session = session.scalar(
                select(AuthSession).where(
                    AuthSession.token_hash == hash_token(token)
                )
            )
        self.assertIsNone(auth_session)
        self.assertEqual(
            self.client.get("/api/auth/me", headers=cookie_headers).status_code,
            401,
        )

    def test_invalid_or_disabled_cookie_session_is_cleared(self):
        invalid = self.client.get(
            "/api/auth/me",
            headers={"Cookie": "delivery_note_session=invalid-token"},
        )
        self.assertEqual(invalid.status_code, 401)
        self.assertIn("Max-Age=0", invalid.headers["set-cookie"])

        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        token = login.json()["token"]
        with self.app.state.database.session() as session:
            session.get(User, 1).active = False
            session.commit()
        disabled = self.client.get(
            "/api/auth/me",
            headers={"Cookie": f"delivery_note_session={token}"},
        )
        self.assertEqual(disabled.status_code, 401)
        self.assertIn("Max-Age=0", disabled.headers["set-cookie"])

    def test_expired_cookie_session_is_cleared(self):
        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        token = login.json()["token"]
        with self.app.state.database.session() as session:
            auth_session = session.scalar(
                select(AuthSession).where(
                    AuthSession.token_hash == hash_token(token)
                )
            )
            auth_session.expires_at = (
                web_api_module.datetime.utcnow()
                - web_api_module.timedelta(seconds=1)
            )
            session.commit()

        expired = self.client.get(
            "/api/auth/me",
            headers={"Cookie": f"delivery_note_session={token}"},
        )
        self.assertEqual(expired.status_code, 401)
        self.assertIn("Max-Age=0", expired.headers["set-cookie"])

    def test_health_endpoints_distinguish_liveness_and_readiness(self):
        for path in ("/health/live", "/health/ready", "/health"):
            response = self.client.get(path)
            self.assertEqual(response.status_code, 200, response.text)
            self.assertEqual(response.json(), {"status": "ok"})

        with patch.object(
            self.app.state.database,
            "session",
            side_effect=RuntimeError("database-secret"),
        ):
            live = self.client.get("/health/live")
            ready = self.client.get("/health/ready")
            legacy = self.client.get("/health")

        self.assertEqual(live.status_code, 200, live.text)
        self.assertEqual(live.json(), {"status": "ok"})
        for response in (ready, legacy):
            self.assertEqual(response.status_code, 503, response.text)
            self.assertNotIn("database-secret", response.text)

    def test_upload_concurrency_and_file_count_limits_must_be_positive(self):
        settings = (
            ("max_concurrent_upload_parses", "MAX_CONCURRENT_UPLOAD_PARSES"),
            ("max_batch_upload_files", "MAX_BATCH_UPLOAD_FILES"),
        )
        for option, message in settings:
            with self.subTest(option=option), TemporaryDirectory() as directory:
                root = Path(directory)
                with self.assertRaisesRegex(ValueError, message):
                    create_app(
                        database_url=f"sqlite+pysqlite:///{root / 'invalid.db'}",
                        storage_root=root / "storage",
                        auto_migrate_schema=False,
                        **{option: 0},
                    )

    def test_upload_parsing_runs_off_loop_and_obeys_concurrency_limit(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            app = create_app(
                database_url=f"sqlite+pysqlite:///{root / 'parse-limit.db'}",
                storage_root=root / "storage",
                bootstrap_admin=("admin", "admin-pass"),
                max_concurrent_upload_parses=1,
            )
            client = SyncASGIClient(app)
            login = client.post(
                "/api/auth/login",
                json={"username": "admin", "password": "admin-pass"},
            )
            self.assertEqual(login.status_code, 200, login.text)
            headers = {"Authorization": f"Bearer {login.json()['token']}"}
            counter_lock = Lock()
            concurrent_parses = Barrier(2)
            active = 0
            maximum_active = 0
            observed_running_loops = []

            def slow_validation(_kind, _path):
                nonlocal active, maximum_active
                try:
                    asyncio.get_running_loop()
                except RuntimeError:
                    observed_running_loops.append(False)
                else:
                    observed_running_loops.append(True)
                with counter_lock:
                    active += 1
                    maximum_active = max(maximum_active, active)
                try:
                    concurrent_parses.wait(timeout=0.3)
                except BrokenBarrierError:
                    pass
                with counter_lock:
                    active -= 1

            async def upload_versions():
                async def keep_event_loop_awake():
                    while True:
                        await asyncio.sleep(0.01)

                heartbeat = asyncio.create_task(keep_event_loop_awake())
                async with AsyncClient(
                    transport=ASGITransport(app=app),
                    base_url="http://testserver",
                ) as async_client:
                    try:
                        return await asyncio.gather(
                            *(
                                async_client.post(
                                    f"/api/input-versions/{kind}",
                                    headers=headers,
                                    data={
                                        "name": f"{kind}-threaded",
                                        "activate": "false",
                                    },
                                    files={
                                        "file": (
                                            f"{kind}.xlsx",
                                            BytesIO(b"content"),
                                        )
                                    },
                                )
                                for kind in ("purchase", "product")
                            )
                        )
                    finally:
                        heartbeat.cancel()
                        with suppress(asyncio.CancelledError):
                            await heartbeat

            try:
                with patch.object(
                    web_api_module,
                    "_validate_input_version",
                    side_effect=slow_validation,
                ):
                    responses = asyncio.run(upload_versions())
            finally:
                client.close()
                app.state.database.dispose()

        self.assertTrue(
            all(response.status_code == 201 for response in responses),
            [response.text for response in responses],
        )
        self.assertEqual(observed_running_loops, [False, False])
        self.assertEqual(maximum_active, 1)

    def test_input_version_activation_keeps_one_active_version(self):
        admin_headers = self.login("admin", "admin-pass")
        created_ids = []
        for name, activate in (("purchase-v1", "true"), ("purchase-v2", "false")):
            response = self.client.post(
                "/api/input-versions/purchase",
                headers=admin_headers,
                data={"name": name, "activate": activate},
                files={
                    "file": (
                        f"{name}.xlsx",
                        BytesIO(self.workbook_bytes("purchase")),
                    )
                },
            )
            self.assertEqual(response.status_code, 201, response.text)
            created_ids.append(response.json()["id"])

        activated = self.client.post(
            f"/api/input-versions/{created_ids[1]}/activate",
            headers=admin_headers,
        )
        self.assertEqual(activated.status_code, 200, activated.text)
        versions = self.client.get("/api/input-versions", headers=admin_headers).json()
        active_ids = [
            version["id"]
            for version in versions
            if version["kind"] == "purchase" and version["active"]
        ]
        self.assertEqual(active_ids, [created_ids[1]])

    def test_input_version_inspection_is_cached_by_version_and_page(self):
        admin_headers = self.login("admin", "admin-pass")
        uploaded = self.client.post(
            "/api/input-versions/product",
            headers=admin_headers,
            data={"name": "product-cache", "activate": "true"},
            files={
                "file": (
                    "product-cache.xlsx",
                    BytesIO(self.workbook_bytes("product")),
                )
            },
        )
        self.assertEqual(uploaded.status_code, 201, uploaded.text)
        version_id = uploaded.json()["id"]

        def inspection_result(kind, _path, offset, limit):
            return {
                "summary": {
                    "kind": kind,
                    "row_count": 1,
                    "columns": ["SKU"],
                    "metrics": {},
                    "issues": [],
                },
                "preview": {
                    "kind": kind,
                    "columns": ["SKU"],
                    "rows": [{"SKU": "SKU-A"}],
                    "total": 1,
                    "offset": offset,
                    "limit": limit,
                },
            }

        with patch.object(
            web_api_module,
            "inspect_input_version_with_preview",
            side_effect=inspection_result,
        ) as inspect:
            summary = self.client.get(
                f"/api/input-versions/{version_id}/summary",
                headers=admin_headers,
            )
            inspection = self.client.get(
                f"/api/input-versions/{version_id}/inspection",
                headers=admin_headers,
            )
            preview = self.client.get(
                f"/api/input-versions/{version_id}/preview",
                headers=admin_headers,
            )
            next_page = self.client.get(
                (f"/api/input-versions/{version_id}/inspection?offset=20&limit=10"),
                headers=admin_headers,
            )

        for response in (summary, inspection, preview, next_page):
            self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(inspect.call_count, 2)
        self.assertEqual(next_page.json()["preview"]["offset"], 20)
        self.assertEqual(next_page.json()["preview"]["limit"], 10)

    @patch.dict(
        "os.environ",
        {
            "GERPGO_API_BASE_URL": "https://example.test",
            "GERPGO_APP_ID": "app",
            "GERPGO_APP_KEY": "key",
        },
    )
    def test_purchase_sync_does_not_require_product_or_supplier_information(self):
        admin_headers = self.login("admin", "admin-pass")
        started = self.client.post(
            "/api/purchase-sync",
            headers=admin_headers,
        )

        self.assertEqual(started.status_code, 201, started.text)
        self.assertIsNone(started.json()["product_version_id"])
        self.assertIsNone(started.json()["supplier_version_id"])

    @patch.dict(
        "os.environ",
        {
            "GERPGO_API_BASE_URL": "https://example.test",
            "GERPGO_APP_ID": "app",
            "GERPGO_APP_KEY": "key",
        },
    )
    def test_operator_can_start_purchase_sync(self):
        admin_headers = self.login("admin", "admin-pass")
        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")

        started = self.client.post(
            "/api/purchase-sync",
            headers=operator_headers,
        )

        self.assertEqual(started.status_code, 201, started.text)
        self.assertEqual(started.json()["status"], "queued")
        status_response = self.client.get(
            "/api/purchase-sync",
            headers=operator_headers,
        )
        self.assertEqual(status_response.status_code, 200, status_response.text)
        self.assertEqual(status_response.json()["job"]["id"], started.json()["id"])

    @patch.dict(
        "os.environ",
        {
            "GERPGO_API_BASE_URL": "https://example.test",
            "GERPGO_APP_ID": "app",
            "GERPGO_APP_KEY": "key",
        },
    )
    def test_operator_can_start_self_operated_inbound_sync(self):
        admin_headers = self.login("admin", "admin-pass")
        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")

        started = self.client.post(
            "/api/self-operated-inbound-sync",
            headers=operator_headers,
        )

        self.assertEqual(started.status_code, 201, started.text)
        self.assertEqual(started.json()["status"], "queued")
        self.assertIsNone(started.json()["base_version_id"])

    def test_admin_can_test_and_save_gerpgo_config(self):
        admin_headers = self.login("admin", "admin-pass")
        payload = {
            "base_url": "https://openapi.example.test/",
            "app_id": "app-001",
            "app_key": "secret-key",
        }

        with patch.object(
            web_api_module.GerpgoClient,
            "authenticate",
        ) as authenticate:
            updated = self.client.put(
                "/api/admin/integrations/gerpgo",
                headers=admin_headers,
                json=payload,
            )

        self.assertEqual(updated.status_code, 200, updated.text)
        self.assertEqual(authenticate.call_count, 1)
        self.assertTrue(updated.json()["configured"])
        self.assertEqual(updated.json()["source"], "managed")
        self.assertEqual(updated.json()["base_url"], "https://openapi.example.test")
        self.assertEqual(updated.json()["app_id_hint"], "ap***01")
        self.assertNotIn("secret-key", updated.text)
        self.assertNotIn("app_key", updated.json())

        status_response = self.client.get(
            "/api/admin/integrations/gerpgo",
            headers=admin_headers,
        )
        self.assertEqual(status_response.status_code, 200, status_response.text)
        self.assertEqual(status_response.json(), updated.json())

        purchase_status = self.client.get(
            "/api/purchase-sync",
            headers=admin_headers,
        )
        self.assertTrue(purchase_status.json()["configured"])

        settings = web_api_module.load_gerpgo_settings(
            self.app.state.storage_root,
        )
        self.assertEqual(settings.app_id, "app-001")
        self.assertEqual(settings.app_key, "secret-key")

        audit_logs = self.client.get(
            "/api/audit-logs",
            headers=admin_headers,
        ).json()
        audit = next(
            item for item in audit_logs if item["action"] == "update_gerpgo_config"
        )
        self.assertNotIn("app_key", audit["details"])
        self.assertNotIn("secret-key", str(audit))

    @patch.dict("os.environ", {}, clear=True)
    def test_unconfigured_gerpgo_uses_official_default_url(self):
        admin_headers = self.login("admin", "admin-pass")

        response = self.client.get(
            "/api/admin/integrations/gerpgo",
            headers=admin_headers,
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertFalse(response.json()["configured"])
        self.assertEqual(response.json()["base_url"], "https://open.gerpgo.com")

    def test_operator_cannot_update_gerpgo_config(self):
        admin_headers = self.login("admin", "admin-pass")
        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")

        response = self.client.put(
            "/api/admin/integrations/gerpgo",
            headers=operator_headers,
            json={
                "base_url": "https://openapi.example.test",
                "app_id": "app",
                "app_key": "key",
            },
        )

        self.assertEqual(response.status_code, 403, response.text)
        self.assertFalse(
            (self.app.state.storage_root / "config" / "gerpgo.json").exists()
        )

    def test_failed_gerpgo_connection_does_not_save_config(self):
        admin_headers = self.login("admin", "admin-pass")
        with patch.object(
            web_api_module.GerpgoClient,
            "authenticate",
            side_effect=web_api_module.GerpgoError("凭证无效"),
        ):
            response = self.client.put(
                "/api/admin/integrations/gerpgo",
                headers=admin_headers,
                json={
                    "base_url": "https://openapi.example.test",
                    "app_id": "wrong-app",
                    "app_key": "wrong-key",
                },
            )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("积加连接验证失败", response.json()["detail"])
        self.assertFalse(
            (self.app.state.storage_root / "config" / "gerpgo.json").exists()
        )

    def test_self_operated_batch_uses_active_api_inbound_version(self):
        headers = self.login("admin", "admin-pass")
        for kind in ("product", "supplier"):
            response = self.client.post(
                f"/api/input-versions/{kind}",
                headers=headers,
                data={"name": f"{kind}-api-batch", "activate": "true"},
                files={
                    "file": (
                        f"{kind}.xlsx",
                        BytesIO(self.workbook_bytes(kind)),
                    )
                },
            )
            self.assertEqual(response.status_code, 201, response.text)

        inbound_path = self.root / "api-inbound.xlsx"
        inbound_path.write_bytes(self.self_operated_inbound_bytes())
        with self.app.state.database.session() as session:
            version = InputVersion(
                kind="self_operated_inbound",
                name="积加待入库-测试版",
                original_name="积加待入库.xlsx",
                storage_path=str(inbound_path),
                active=True,
                created_by=1,
            )
            session.add(version)
            session.commit()
            inbound_version_id = version.id

        created = self.client.post(
            "/api/self-operated-batches",
            headers=headers,
            data={"name": "使用 API 待入库数据的批次"},
            files={
                "delivery_file": (
                    "质检交货单.xlsx",
                    BytesIO(self.self_operated_delivery_bytes()),
                )
            },
        )

        self.assertEqual(created.status_code, 201, created.text)
        batch = created.json()
        self.assertEqual(
            batch["version_ids"]["self_operated_inbound"],
            inbound_version_id,
        )
        self.assertEqual(
            batch["versions"]["self_operated_inbound"]["name"],
            "积加待入库-测试版",
        )
        self.assertEqual(
            batch["inbound_file"]["original_name"],
            "积加待入库.xlsx",
        )

    def test_purchase_sync_issues_can_be_previewed_by_an_operator(self):
        admin_headers = self.login("admin", "admin-pass")
        with self.app.state.database.session() as session:
            job = PurchaseSyncJob(
                status="succeeded",
                created_by=1,
                issues=[
                    {
                        "severity": "warning",
                        "message": "共享站点数据不能参与正常交货匹配",
                        "po_code": "PO-1001",
                        "sku": "SKU-A",
                        "source_site": "共享",
                        "supplier_code": "SUP-1",
                        "supplier_name": "供应商 A",
                        "warehouse": "水鞋-广州仓",
                        "quantity": 12,
                        "code": "shared_site",
                    }
                ],
            )
            session.add(job)
            session.commit()
            job_id = job.id

        preview = self.client.get(
            f"/api/purchase-sync/{job_id}/issues",
            headers=admin_headers,
        )

        self.assertEqual(preview.status_code, 200, preview.text)
        self.assertEqual(preview.json()[0]["po_code"], "PO-1001")
        self.assertEqual(preview.json()[0]["warehouse"], "水鞋-广州仓")
        self.assertEqual(preview.json()[0]["quantity"], 12)
        download = self.client.get(
            f"/api/purchase-sync/{job_id}/issues/download",
            headers=admin_headers,
        )
        self.assertEqual(download.status_code, 200, download.text)
        issue_frame = pd.read_excel(BytesIO(download.content))
        self.assertEqual(issue_frame.loc[0, "目的仓"], "水鞋-广州仓")
        self.assertEqual(issue_frame.loc[0, "未交量"], 12)
        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")
        operator_preview = self.client.get(
            f"/api/purchase-sync/{job_id}/issues",
            headers=operator_headers,
        )
        self.assertEqual(operator_preview.status_code, 200, operator_preview.text)
        self.assertEqual(operator_preview.json()[0]["po_code"], "PO-1001")

    def test_purchase_sync_candidate_can_be_previewed_by_an_operator(self):
        admin_headers = self.login("admin", "admin-pass")
        candidate_path = self.root / "purchase-candidate.xlsx"
        candidate_path.write_bytes(self.workbook_bytes("purchase"))
        with self.app.state.database.session() as session:
            version = InputVersion(
                kind="purchase",
                name="采购候选版本",
                original_name="purchase-candidate.xlsx",
                storage_path=str(candidate_path),
                active=False,
                created_by=1,
            )
            session.add(version)
            session.flush()
            job = PurchaseSyncJob(
                status="succeeded",
                created_by=1,
                candidate_version_id=version.id,
            )
            session.add(job)
            session.commit()
            job_id = job.id

        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")
        preview = self.client.get(
            f"/api/purchase-sync/{job_id}/preview?limit=100",
            headers=operator_headers,
        )

        self.assertEqual(preview.status_code, 200, preview.text)
        payload = preview.json()
        self.assertEqual(
            payload["columns"],
            ["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"],
        )
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["rows"][0]["_row_number"], 1)
        self.assertEqual(payload["rows"][0]["SKU"], "SKU-A")
        self.assertEqual(payload["rows"][0]["目的仓"], "水鞋-广州仓")
        self.assertEqual(payload["rows"][0]["未交量"], 100)

    def test_self_operated_candidate_preview_has_stable_row_numbers(self):
        admin_headers = self.login("admin", "admin-pass")
        candidate_path = self.root / "self-operated-candidate.xlsx"
        candidate_path.write_bytes(self.self_operated_inbound_bytes())
        with self.app.state.database.session() as session:
            version = InputVersion(
                kind="self_operated_inbound",
                name="待入库候选版本",
                original_name="self-operated-candidate.xlsx",
                storage_path=str(candidate_path),
                active=False,
                created_by=1,
            )
            session.add(version)
            session.flush()
            job = SelfOperatedInboundSyncJob(
                status="succeeded",
                created_by=1,
                candidate_version_id=version.id,
                issues=[
                    {
                        "severity": "warning",
                        "message": "共享站点数据不能参与正常入库匹配",
                        "order_no": "IN-1",
                        "sku": "SKU-A",
                        "source_site": "AMAZON:SEEKWAY:US",
                        "supplier_code": "GYS-023",
                        "supplier_name": "KuangBiao",
                        "code": "shared_site",
                    }
                ],
            )
            session.add(job)
            session.commit()
            job_id = job.id

        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")
        preview = self.client.get(
            f"/api/self-operated-inbound-sync/{job_id}/preview?limit=100",
            headers=operator_headers,
        )

        self.assertEqual(preview.status_code, 200, preview.text)
        payload = preview.json()
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["rows"][0]["_row_number"], 1)
        self.assertEqual(payload["rows"][0]["入库单号"], "IN-1")
        self.assertEqual(payload["rows"][0]["SKU"], "SKU-A")
        self.assertEqual(payload["rows"][0]["应收货"], 10)

        issues = self.client.get(
            f"/api/self-operated-inbound-sync/{job_id}/issues",
            headers=operator_headers,
        )
        self.assertEqual(issues.status_code, 200, issues.text)
        issue = issues.json()[0]
        self.assertEqual(issue["warehouse"], "自营仓")
        self.assertEqual(issue["remaining_quantity"], 10)
        self.assertEqual(issue["purchase_code"], "PO-20260801")
        self.assertEqual(issue["related_code"], "LN2608179025")

        download = self.client.get(
            f"/api/self-operated-inbound-sync/{job_id}/issues/download",
            headers=operator_headers,
        )
        self.assertEqual(download.status_code, 200, download.text)
        issue_frame = pd.read_excel(BytesIO(download.content))
        self.assertEqual(issue_frame.loc[0, "入库仓"], "自营仓")
        self.assertEqual(issue_frame.loc[0, "剩余应收货"], 10)

    def test_initial_state_includes_builtin_templates(self):
        admin_headers = self.login("admin", "admin-pass")
        versions = self.client.get(
            "/api/input-versions",
            headers=admin_headers,
        ).json()
        export_versions = [
            version for version in versions if version["kind"] == "template"
        ]
        inbound_versions = [
            version for version in versions if version["kind"] == "inbound_template"
        ]

        self.assertEqual(len(export_versions), 1)
        export_version = export_versions[0]
        self.assertEqual(
            export_version["name"],
            "系统内置交货导出模板",
        )
        self.assertEqual(
            export_version["original_name"],
            "交货导入模板.xlsx",
        )
        self.assertTrue(export_version["active"])

        export_download = self.client.get(
            f"/api/input-versions/{export_version['id']}/download",
            headers=admin_headers,
        )
        self.assertEqual(export_download.status_code, 200, export_download.text)
        export_workbook = load_workbook(
            BytesIO(export_download.content),
            read_only=True,
        )
        try:
            self.assertEqual(
                [
                    export_workbook.active.cell(row=2, column=column).value
                    for column in range(1, len(IMPORT_COLUMNS) + 1)
                ],
                IMPORT_COLUMNS,
            )
            self.assertEqual(export_workbook.active["A3"].value, "示例仓库")
        finally:
            export_workbook.close()

        self.assertEqual(len(inbound_versions), 1)
        version = inbound_versions[0]
        self.assertEqual(version["name"], "系统内置积加入库模板")
        self.assertEqual(version["original_name"], "积加批量入库模板.xlsx")
        self.assertTrue(version["active"])

        downloaded = self.client.get(
            f"/api/input-versions/{version['id']}/download",
            headers=admin_headers,
        )
        self.assertEqual(downloaded.status_code, 200, downloaded.text)
        workbook = load_workbook(BytesIO(downloaded.content), read_only=True)
        try:
            headers = [
                workbook.active.cell(row=1, column=column).value
                for column in range(1, len(INBOUND_TEMPLATE_COLUMNS) + 1)
            ]
            self.assertEqual(headers, INBOUND_TEMPLATE_COLUMNS)
            self.assertEqual(
                sum(
                    1 for sheet in workbook.worksheets if sheet.sheet_state == "hidden"
                ),
                70,
            )
        finally:
            workbook.close()

    def test_delivery_batch_uses_builtin_template_without_upload(self):
        admin_headers = self.login("admin", "admin-pass")
        for kind in ("purchase", "product", "supplier", "position"):
            response = self.client.post(
                f"/api/input-versions/{kind}",
                headers=admin_headers,
                data={"name": f"{kind}-v1", "activate": "true"},
                files={
                    "file": (
                        f"{kind}.xlsx",
                        BytesIO(self.workbook_bytes(kind)),
                    )
                },
            )
            self.assertEqual(response.status_code, 201, response.text)

        created = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "使用系统内置导出模板"},
        )

        self.assertEqual(created.status_code, 201, created.text)
        self.assertEqual(
            created.json()["versions"]["template"]["name"],
            "系统内置交货导出模板",
        )

    def test_self_operated_batch_only_requires_its_own_input_versions(self):
        headers = self.login("admin", "admin-pass")
        version_ids = {}
        for kind in ("product", "supplier"):
            response = self.client.post(
                f"/api/input-versions/{kind}",
                headers=headers,
                data={"name": f"{kind}-self-operated", "activate": "true"},
                files={
                    "file": (
                        f"{kind}.xlsx",
                        BytesIO(self.workbook_bytes(kind)),
                    )
                },
            )
            self.assertEqual(response.status_code, 201, response.text)
            version_ids[kind] = response.json()["id"]

        created = self.client.post(
            "/api/self-operated-batches",
            headers=headers,
            data={"name": "不依赖原流程资料的自营仓批次"},
            files={
                "delivery_file": (
                    "质检交货单.xlsx",
                    BytesIO(self.self_operated_delivery_bytes()),
                ),
                "inbound_file": (
                    "自营仓收货入库单.xlsx",
                    BytesIO(self.self_operated_inbound_bytes()),
                ),
            },
        )

        self.assertEqual(created.status_code, 201, created.text)
        batch = created.json()
        self.assertEqual(batch["version_ids"]["product"], version_ids["product"])
        self.assertEqual(batch["version_ids"]["supplier"], version_ids["supplier"])
        self.assertIsNone(batch["version_ids"]["purchase"])
        self.assertIsNone(batch["version_ids"]["position"])
        self.assertIsNone(batch["version_ids"]["template"])
        self.assertIn("inbound_template", batch["versions"])
        self.assertEqual(batch["file_count"], 1)
        self.assertTrue(batch["inbound_file"]["uploaded"])

        delivery_batch = self.client.post(
            "/api/batches",
            headers=headers,
            json={"name": "原流程仍要求完整资料"},
        )
        self.assertEqual(delivery_batch.status_code, 409, delivery_batch.text)
        self.assertIn("purchase", delivery_batch.json()["detail"])
        self.assertIn("position", delivery_batch.json()["detail"])
        self.assertNotIn("template", delivery_batch.json()["detail"])

    def test_position_bootstrap_upload_is_allowed(self):
        admin_headers = self.login("admin", "admin-pass")

        response = self.client.post(
            "/api/input-versions/position",
            headers=admin_headers,
            data={"name": "position-bootstrap", "activate": "true"},
            files={
                "file": (
                    "position-bootstrap.xlsx",
                    BytesIO(self.workbook_bytes("position")),
                )
            },
        )

        self.assertEqual(response.status_code, 201, response.text)
        self.assertTrue(response.json()["active"])

    def test_invalid_input_version_is_rejected_before_activation(self):
        admin_headers = self.login("admin", "admin-pass")
        response = self.client.post(
            "/api/input-versions/purchase",
            headers=admin_headers,
            data={"name": "broken-purchase", "activate": "true"},
            files={"file": ("broken.xlsx", BytesIO(b"not-an-excel-file"))},
        )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("输入版本校验失败", response.json()["detail"])
        versions = self.client.get("/api/input-versions", headers=admin_headers).json()
        self.assertEqual(
            [version["kind"] for version in versions],
            ["inbound_template", "template"],
        )
        purchase_root = self.app.state.storage_root / "master" / "purchase"
        self.assertEqual(list(purchase_root.glob("*")), [])

    def test_initial_state_creates_batches_without_overreceipt_rule(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)

        rules = self.client.get(
            "/api/overreceipt-rule-versions",
            headers=admin_headers,
        )
        self.assertEqual(rules.status_code, 200, rules.text)
        self.assertEqual(rules.json(), [])

        created = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "默认关闭超收"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        self.assertIsNone(created.json()["overreceipt_rule"])

        logs = self.client.get("/api/audit-logs", headers=admin_headers).json()
        create_log = next(
            log
            for log in logs
            if log["action"] == "create_batch"
            and log["entity_id"] == str(created.json()["id"])
        )
        self.assertIsNone(create_log["details"]["overreceipt_rule_version_id"])

    def test_overreceipt_warehouses_cache_the_active_purchase_version(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        original_reader = web_api_module.read_purchase_workbook

        with patch.object(
            web_api_module,
            "read_purchase_workbook",
            wraps=original_reader,
        ) as reader:
            first = self.client.get(
                "/api/overreceipt-rule-versions/warehouses",
                headers=admin_headers,
            )
            second = self.client.get(
                "/api/overreceipt-rule-versions/warehouses",
                headers=admin_headers,
            )

        self.assertEqual(first.status_code, 200, first.text)
        self.assertEqual(second.status_code, 200, second.text)
        self.assertEqual(first.json(), ["水鞋-广州仓"])
        self.assertEqual(second.json(), first.json())
        self.assertEqual(reader.call_count, 1)

        uploaded = self.client.post(
            "/api/input-versions/purchase",
            headers=admin_headers,
            data={"name": "purchase-v2", "activate": "true"},
            files={
                "file": (
                    "purchase-v2.xlsx",
                    BytesIO(self.workbook_bytes("purchase")),
                )
            },
        )
        self.assertEqual(uploaded.status_code, 201, uploaded.text)

        with patch.object(
            web_api_module,
            "read_purchase_workbook",
            wraps=original_reader,
        ) as reader:
            third = self.client.get(
                "/api/overreceipt-rule-versions/warehouses",
                headers=admin_headers,
            )

        self.assertEqual(third.status_code, 200, third.text)
        self.assertEqual(third.json(), first.json())
        self.assertEqual(reader.call_count, 1)

    def test_operator_can_publish_activate_and_lock_immutable_overreceipt_rules(self):
        admin_headers = self.login("admin", "admin-pass")
        operator = self.create_operator(admin_headers)
        self.upload_active_versions(admin_headers)
        operator_headers = self.login("operator", "operator-pass")

        warehouses = self.client.get(
            "/api/overreceipt-rule-versions/warehouses",
            headers=operator_headers,
        )
        self.assertEqual(warehouses.status_code, 200, warehouses.text)
        self.assertEqual(warehouses.json(), ["水鞋-广州仓"])

        first = self.client.post(
            "/api/overreceipt-rule-versions",
            headers=operator_headers,
            json={
                "name": "2026-07 短尾放宽",
                "short_tail_limit": 50,
                "medium_tail_limit": 20,
                "long_tail_limit": 10,
                "allowed_warehouses": ["水鞋-广州仓"],
            },
        )
        self.assertEqual(first.status_code, 201, first.text)
        self.assertTrue(first.json()["active"])
        self.assertEqual(first.json()["created_by"], operator["id"])

        locked_batch = self.client.post(
            "/api/batches",
            headers=operator_headers,
            json={"name": "锁定超收规则 V1"},
        )
        self.assertEqual(locked_batch.status_code, 201, locked_batch.text)
        self.assertEqual(
            locked_batch.json()["overreceipt_rule"]["id"],
            first.json()["id"],
        )

        second = self.client.post(
            "/api/overreceipt-rule-versions",
            headers=admin_headers,
            json={
                "name": "2026-08 收紧",
                "short_tail_limit": 30,
                "medium_tail_limit": 10,
                "long_tail_limit": 0,
                "allowed_warehouses": [],
            },
        )
        self.assertEqual(second.status_code, 201, second.text)
        self.assertTrue(second.json()["active"])

        versions = self.client.get(
            "/api/overreceipt-rule-versions",
            headers=operator_headers,
        )
        self.assertEqual(versions.status_code, 200, versions.text)
        by_id = {item["id"]: item for item in versions.json()}
        self.assertFalse(by_id[first.json()["id"]]["active"])
        self.assertTrue(by_id[second.json()["id"]]["active"])

        reactivated = self.client.post(
            f"/api/overreceipt-rule-versions/{first.json()['id']}/activate",
            headers=operator_headers,
        )
        self.assertEqual(reactivated.status_code, 200, reactivated.text)
        self.assertTrue(reactivated.json()["active"])

        unchanged_batch = self.client.get(
            f"/api/batches/{locked_batch.json()['id']}",
            headers=operator_headers,
        )
        self.assertEqual(
            unchanged_batch.json()["overreceipt_rule"]["id"],
            first.json()["id"],
        )

        logs = self.client.get("/api/audit-logs", headers=admin_headers).json()
        actions = [log for log in logs if log["entity_type"] == "overreceipt_rule"]
        self.assertEqual(
            [log["action"] for log in actions],
            [
                "activate_overreceipt_rule",
                "publish_overreceipt_rule",
                "publish_overreceipt_rule",
            ],
        )
        self.assertEqual(actions[0]["user_id"], operator["id"])

    def test_rule_version_names_can_change_without_changing_locked_rules(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)

        first = self.client.post(
            "/api/overreceipt-rule-versions",
            headers=headers,
            json={
                "name": "交货超收旧名称",
                "short_tail_limit": 50,
                "medium_tail_limit": 20,
                "long_tail_limit": 10,
                "allowed_warehouses": ["水鞋-广州仓"],
            },
        )
        self.assertEqual(first.status_code, 201, first.text)
        locked_batch = self.client.post(
            "/api/batches",
            headers=headers,
            json={"name": "规则名称修改验收"},
        )
        self.assertEqual(locked_batch.status_code, 201, locked_batch.text)
        second = self.client.post(
            "/api/overreceipt-rule-versions",
            headers=headers,
            json={
                "name": "交货超收当前名称",
                "short_tail_limit": 30,
                "medium_tail_limit": 10,
                "long_tail_limit": 0,
                "allowed_warehouses": [],
            },
        )
        self.assertEqual(second.status_code, 201, second.text)

        renamed = self.client.put(
            f"/api/overreceipt-rule-versions/{first.json()['id']}/name",
            headers=headers,
            json={"name": " 交货超收新名称 "},
        )
        self.assertEqual(renamed.status_code, 200, renamed.text)
        self.assertEqual(renamed.json()["id"], first.json()["id"])
        self.assertEqual(renamed.json()["name"], "交货超收新名称")
        self.assertEqual(renamed.json()["short_tail_limit"], 50)
        self.assertEqual(renamed.json()["medium_tail_limit"], 20)
        self.assertEqual(renamed.json()["long_tail_limit"], 10)
        self.assertEqual(
            renamed.json()["allowed_warehouses"],
            ["水鞋-广州仓"],
        )
        self.assertFalse(renamed.json()["active"])

        locked = self.client.get(
            f"/api/batches/{locked_batch.json()['id']}",
            headers=headers,
        )
        self.assertEqual(locked.status_code, 200, locked.text)
        self.assertEqual(
            locked.json()["overreceipt_rule"]["id"],
            first.json()["id"],
        )
        self.assertEqual(
            locked.json()["overreceipt_rule"]["name"],
            "交货超收新名称",
        )
        self.assertEqual(
            locked.json()["overreceipt_rule"]["short_tail_limit"],
            50,
        )

        duplicate = self.client.put(
            f"/api/overreceipt-rule-versions/{first.json()['id']}/name",
            headers=headers,
            json={"name": second.json()["name"]},
        )
        self.assertEqual(duplicate.status_code, 409, duplicate.text)
        blank = self.client.put(
            f"/api/overreceipt-rule-versions/{first.json()['id']}/name",
            headers=headers,
            json={"name": "   "},
        )
        self.assertEqual(blank.status_code, 400, blank.text)

        self_operated = self.client.post(
            "/api/self-operated-overreceipt-rule-versions",
            headers=headers,
            json={"name": "自营仓旧名称", "allowance": 5},
        )
        self.assertEqual(self_operated.status_code, 201, self_operated.text)
        renamed_self_operated = self.client.put(
            "/api/self-operated-overreceipt-rule-versions/"
            f"{self_operated.json()['id']}/name",
            headers=headers,
            json={"name": "自营仓新名称"},
        )
        self.assertEqual(
            renamed_self_operated.status_code,
            200,
            renamed_self_operated.text,
        )
        self.assertEqual(
            renamed_self_operated.json()["id"],
            self_operated.json()["id"],
        )
        self.assertEqual(
            renamed_self_operated.json()["name"],
            "自营仓新名称",
        )
        self.assertEqual(renamed_self_operated.json()["allowance"], 5)

        logs = self.client.get("/api/audit-logs", headers=headers).json()
        rename_actions = {
            log["action"]: log for log in logs if log["action"].startswith("rename_")
        }
        self.assertEqual(
            set(rename_actions),
            {
                "rename_overreceipt_rule",
                "rename_self_operated_overreceipt_rule",
            },
        )
        self.assertEqual(
            rename_actions["rename_overreceipt_rule"]["details"],
            {"before": "交货超收旧名称", "after": "交货超收新名称"},
        )

    def test_self_operated_batch_locks_rule_and_creates_with_two_business_files(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)
        template = self.client.post(
            "/api/input-versions/inbound_template",
            headers=headers,
            data={"name": "inbound-template-v1", "activate": "true"},
            files={
                "file": (
                    "inbound-template.xlsx",
                    BytesIO(self.workbook_bytes("inbound_template")),
                )
            },
        )
        self.assertEqual(template.status_code, 201, template.text)
        rule = self.client.post(
            "/api/self-operated-overreceipt-rule-versions",
            headers=headers,
            json={"name": "自营仓超收 5 件", "allowance": 5},
        )
        self.assertEqual(rule.status_code, 201, rule.text)

        created = self.client.post(
            "/api/self-operated-batches",
            headers=headers,
            data={"name": "自营仓入库接口测试"},
            files={
                "delivery_file": (
                    "260817-狂飙-质检交货单.xlsx",
                    BytesIO(self.self_operated_delivery_bytes()),
                ),
                "inbound_file": (
                    "自营仓收货入库单.xlsx",
                    BytesIO(self.self_operated_inbound_bytes()),
                ),
            },
        )
        self.assertEqual(created.status_code, 201, created.text)
        batch = created.json()
        batch_id = batch["id"]
        self.assertEqual(batch["workflow"], "self_operated_inbound")
        self.assertEqual(batch["self_operated_overreceipt_rule"]["allowance"], 5)
        self.assertEqual(
            batch["versions"]["inbound_template"]["id"],
            template.json()["id"],
        )
        self.assertEqual(batch["file_count"], 1)
        self.assertTrue(batch["inbound_file"]["uploaded"])
        listed_batch = next(
            item
            for item in self.client.get("/api/batches", headers=headers).json()
            if item["id"] == batch_id
        )
        self.assertEqual(
            listed_batch,
            {key: batch[key] for key in listed_batch},
        )
        self.app.state.max_batch_upload_files = 1
        extra_file = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=headers,
            files={"file": ("额外质检交货单.xlsx", BytesIO(self.delivery_bytes()))},
        )
        self.assertEqual(extra_file.status_code, 409, extra_file.text)
        self.assertIn("只能上传一份质检交货单", extra_file.json()["detail"])
        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=headers,
        )
        self.assertEqual(preflight.status_code, 200, preflight.text)
        self.assertEqual(preflight.json()["status"], "preflight_ready")

    def test_self_operated_creation_is_atomic_when_file_validation_fails(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)

        invalid = self.client.post(
            "/api/self-operated-batches",
            headers=headers,
            data={"name": "校验失败的自营仓批次"},
            files={
                "delivery_file": (
                    "质检交货单.xlsx",
                    BytesIO(self.self_operated_delivery_bytes()),
                ),
                "inbound_file": ("异常入库单.xlsx", BytesIO(b"not-an-excel-file")),
            },
        )
        self.assertEqual(invalid.status_code, 400, invalid.text)

        batches = self.client.get("/api/batches", headers=headers)
        self.assertEqual(batches.status_code, 200, batches.text)
        self.assertEqual(batches.json(), [])
        temporary_root = (
            self.app.state.storage_root / "temporary" / "self-operated-batches"
        )
        self.assertEqual(list(temporary_root.glob("*")), [])

    def test_delivery_creation_with_files_is_atomic(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)

        invalid = self.client.post(
            "/api/batches/with-files",
            headers=headers,
            data={"name": "校验失败的交货批次"},
            files={"files": ("异常交货单.xlsx", BytesIO(b"not-an-excel-file"))},
        )
        self.assertEqual(invalid.status_code, 400, invalid.text)
        self.assertEqual(
            self.client.get("/api/batches", headers=headers).json(),
            [],
        )
        temporary_root = (
            self.app.state.storage_root / "temporary" / "delivery-batches"
        )
        self.assertEqual(list(temporary_root.glob("*")), [])

        created = self.client.post(
            "/api/batches/with-files",
            headers=headers,
            data={"name": "带文件创建的交货批次"},
            files={"files": ("交货单.xlsx", BytesIO(self.delivery_bytes()))},
        )
        self.assertEqual(created.status_code, 201, created.text)
        self.assertEqual(created.json()["file_count"], 1)

    def test_delivery_creation_rejects_too_many_files_before_writing(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)
        self.app.state.max_batch_upload_files = 1

        response = self.client.post(
            "/api/batches/with-files",
            headers=headers,
            data={"name": "文件过多"},
            files=[
                ("files", ("one.xlsx", BytesIO(self.delivery_bytes()))),
                ("files", ("two.xlsx", BytesIO(self.delivery_bytes()))),
            ],
        )

        self.assertEqual(response.status_code, 413, response.text)
        self.assertIn("最多上传 1 份", response.json()["detail"])
        self.assertEqual(
            self.client.get("/api/batches", headers=headers).json(),
            [],
        )
        temporary_root = (
            self.app.state.storage_root / "temporary" / "delivery-batches"
        )
        self.assertFalse(temporary_root.exists())

    def test_empty_delivery_drafts_are_removed_without_touching_uploaded_batches(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)
        empty = self.client.post(
            "/api/batches",
            headers=headers,
            json={"name": "应被清理的空交货批次"},
        )
        self.assertEqual(empty.status_code, 201, empty.text)
        created = self.client.post(
            "/api/batches/with-files",
            headers=headers,
            data={"name": "保留的交货批次"},
            files={"files": ("交货单.xlsx", BytesIO(self.delivery_bytes()))},
        )
        self.assertEqual(created.status_code, 201, created.text)

        cleaned = self.client.delete("/api/batches/empty", headers=headers)
        self.assertEqual(cleaned.status_code, 200, cleaned.text)
        self.assertEqual(cleaned.json()["deleted_ids"], [empty.json()["id"]])
        batches = self.client.get("/api/batches", headers=headers)
        self.assertEqual(
            [batch["name"] for batch in batches.json()],
            ["保留的交货批次"],
        )

    def test_empty_self_operated_drafts_are_removed_without_touching_ready_batches(
        self,
    ):
        headers = self.login("admin", "admin-pass")
        version_ids = self.upload_active_versions(headers)
        created = self.client.post(
            "/api/self-operated-batches",
            headers=headers,
            data={"name": "保留的自营仓批次"},
            files={
                "delivery_file": (
                    "质检交货单.xlsx",
                    BytesIO(self.self_operated_delivery_bytes()),
                ),
                "inbound_file": (
                    "自营仓收货入库单.xlsx",
                    BytesIO(self.self_operated_inbound_bytes()),
                ),
            },
        )
        self.assertEqual(created.status_code, 201, created.text)

        with self.app.state.database.session() as session:
            empty = Batch(
                name="应被清理的空批次",
                created_by=1,
                purchase_version_id=None,
                product_version_id=version_ids["product"],
                supplier_version_id=version_ids["supplier"],
                position_version_id=None,
                template_version_id=None,
            )
            session.add(empty)
            session.flush()
            session.add(
                SelfOperatedBatch(
                    batch_id=empty.id,
                    template_version_id=created.json()["versions"]["inbound_template"][
                        "id"
                    ],
                )
            )
            session.commit()
            empty_id = empty.id

        cleaned = self.client.delete(
            "/api/self-operated-batches/empty",
            headers=headers,
        )
        self.assertEqual(cleaned.status_code, 200, cleaned.text)
        self.assertEqual(cleaned.json()["deleted_ids"], [empty_id])

        batches = self.client.get("/api/batches", headers=headers)
        self.assertEqual(batches.status_code, 200, batches.text)
        self.assertEqual(
            [batch["name"] for batch in batches.json()], ["保留的自营仓批次"]
        )

    def test_admin_can_delete_multiple_batches_and_owned_files(self):
        headers = self.login("admin", "admin-pass")
        self.upload_active_versions(headers)
        delivery = self.client.post(
            "/api/batches/with-files",
            headers=headers,
            data={"name": "delete-delivery-batch"},
            files={"files": ("delivery.xlsx", BytesIO(self.delivery_bytes()))},
        )
        self.assertEqual(delivery.status_code, 201, delivery.text)
        self_operated = self.client.post(
            "/api/self-operated-batches",
            headers=headers,
            data={"name": "delete-self-operated-batch"},
            files={
                "delivery_file": (
                    "quality-delivery.xlsx",
                    BytesIO(self.self_operated_delivery_bytes()),
                ),
                "inbound_file": (
                    "self-operated-inbound.xlsx",
                    BytesIO(self.self_operated_inbound_bytes()),
                ),
            },
        )
        self.assertEqual(self_operated.status_code, 201, self_operated.text)
        batch_ids = [delivery.json()["id"], self_operated.json()["id"]]
        batch_directories = [
            self.root / "storage" / "batches" / str(batch_id) for batch_id in batch_ids
        ]
        self.assertTrue(all(path.is_dir() for path in batch_directories))

        with self.app.state.database.session() as session:
            source = session.scalar(
                select(BatchFile).where(BatchFile.batch_id == batch_ids[0])
            )
            exception = ExceptionRecord(
                batch_file_id=source.id,
                sku="SKU-A",
                delivery_quantity=1,
                allocated_quantity=0,
                manual_quantity=1,
                reason="delete-test",
            )
            session.add(exception)
            session.flush()
            session.add(SplitRecord(exception_id=exception.id, quantity=1))
            session.add(
                Job(
                    batch_id=batch_ids[0],
                    kind="compute",
                    status="succeeded",
                )
            )
            session.commit()

        deleted = self.client.request(
            "DELETE",
            "/api/batches",
            headers=headers,
            json={"batch_ids": batch_ids},
        )

        self.assertEqual(deleted.status_code, 200, deleted.text)
        self.assertEqual(deleted.json()["deleted_ids"], batch_ids)
        self.assertEqual(deleted.json()["file_cleanup_failed_ids"], [])
        self.assertTrue(all(not path.exists() for path in batch_directories))
        self.assertEqual(
            self.client.get("/api/batches", headers=headers).json(),
            [],
        )

    def test_batch_delete_is_admin_only_and_rejects_active_batch_atomically(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        first = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "keep-batch"},
        ).json()
        active = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "running-batch"},
        ).json()
        with self.app.state.database.session() as session:
            session.get(Batch, active["id"]).status = "running"
            session.commit()

        operator = self.create_operator(admin_headers)
        operator_headers = self.login(operator["username"], "operator-pass")
        forbidden = self.client.request(
            "DELETE",
            "/api/batches",
            headers=operator_headers,
            json={"batch_ids": [first["id"]]},
        )
        self.assertEqual(forbidden.status_code, 403, forbidden.text)

        blocked = self.client.request(
            "DELETE",
            "/api/batches",
            headers=admin_headers,
            json={"batch_ids": [first["id"], active["id"]]},
        )
        self.assertEqual(blocked.status_code, 409, blocked.text)
        self.assertIn("running-batch", blocked.json()["detail"])
        remaining_ids = [
            batch["id"]
            for batch in self.client.get(
                "/api/batches",
                headers=admin_headers,
            ).json()
        ]
        self.assertEqual(set(remaining_ids), {first["id"], active["id"]})

    def test_admin_can_disable_and_reset_operator_password(self):
        admin_headers = self.login("admin", "admin-pass")
        operator = self.create_operator(admin_headers)

        disabled = self.client.put(
            f"/api/users/{operator['id']}/status",
            headers=admin_headers,
            json={"active": False},
        )
        self.assertEqual(disabled.status_code, 200, disabled.text)
        self.assertFalse(disabled.json()["active"])
        self.assertEqual(
            self.client.post(
                "/api/auth/login",
                json={"username": "operator", "password": "operator-pass"},
            ).status_code,
            401,
        )

        self_deactivate = self.client.put(
            "/api/users/1/status",
            headers=admin_headers,
            json={"active": False},
        )
        self.assertEqual(self_deactivate.status_code, 409)

        enabled = self.client.put(
            f"/api/users/{operator['id']}/status",
            headers=admin_headers,
            json={"active": True},
        )
        self.assertTrue(enabled.json()["active"])
        reset = self.client.put(
            f"/api/users/{operator['id']}/password",
            headers=admin_headers,
            json={"password": "operator-new-pass"},
        )
        self.assertEqual(reset.status_code, 204, reset.text)
        self.assertEqual(
            self.client.post(
                "/api/auth/login",
                json={"username": "operator", "password": "operator-pass"},
            ).status_code,
            401,
        )
        self.login("operator", "operator-new-pass")

    def test_api_timestamps_include_an_explicit_utc_offset(self):
        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        self.assertEqual(login.status_code, 200, login.text)
        self.assertTrue(login.json()["expires_at"].endswith("Z"))
        headers = {"Authorization": f"Bearer {login.json()['token']}"}

        self.upload_active_versions(headers)
        versions = self.client.get(
            "/api/input-versions",
            headers=headers,
        ).json()
        self.assertTrue(
            all(version["created_at"].endswith("Z") for version in versions)
        )

        created = self.client.post(
            "/api/batches",
            headers=headers,
            json={"name": "北京时间边界测试"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        self.assertTrue(created.json()["created_at"].endswith("Z"))
        self.assertTrue(created.json()["updated_at"].endswith("Z"))

        logs = self.client.get("/api/audit-logs", headers=headers).json()
        self.assertTrue(all(log["created_at"].endswith("Z") for log in logs))

    def test_versions_batch_order_preflight_and_compute_job(self):
        admin_headers = self.login("admin", "admin-pass")
        self.create_operator(admin_headers)
        version_ids = self.upload_active_versions(admin_headers)
        operator_headers = self.login("operator", "operator-pass")

        created = self.client.post(
            "/api/batches",
            headers=operator_headers,
            json={"name": "7 月交货批次"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        batch = created.json()
        self.assertEqual(batch["version_ids"], version_ids)
        self.assertEqual(
            {kind: item["name"] for kind, item in batch["versions"].items()},
            {kind: f"{kind}-v1" for kind in INPUT_KINDS},
        )
        self.assertEqual(batch["jobs"], {})
        batch_id = batch["id"]

        first = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=operator_headers,
            files={
                "file": (
                    "260717-狂飙-A交货单-发货10箱.xlsx",
                    BytesIO(self.delivery_bytes()),
                )
            },
        )
        duplicate = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=operator_headers,
            files={
                "file": (
                    "260717-狂飙-A交货单-发货10箱.xlsx",
                    BytesIO(self.delivery_bytes()),
                )
            },
        )
        second = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=operator_headers,
            files={
                "file": (
                    "260717-狂飙-B交货单-发货20箱.xlsx",
                    BytesIO(self.delivery_bytes()),
                )
            },
        )
        self.assertEqual(first.status_code, 201, first.text)
        self.assertEqual(duplicate.status_code, 409, duplicate.text)
        self.assertEqual(second.status_code, 201, second.text)

        reordered = self.client.put(
            f"/api/batches/{batch_id}/files/order",
            headers=operator_headers,
            json={"file_ids": [second.json()["id"], first.json()["id"]]},
        )
        self.assertEqual(reordered.status_code, 200, reordered.text)
        self.assertEqual(
            [item["id"] for item in reordered.json()["files"]],
            [second.json()["id"], first.json()["id"]],
        )

        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=operator_headers,
        )
        self.assertEqual(preflight.status_code, 200, preflight.text)
        self.assertEqual(preflight.json()["status"], "preflight_ready")

        first_start = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=operator_headers,
        )
        second_start = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=operator_headers,
        )
        self.assertEqual(first_start.status_code, 202, first_start.text)
        self.assertEqual(second_start.status_code, 202, second_start.text)
        self.assertEqual(first_start.json()["id"], second_start.json()["id"])
        self.assertEqual(first_start.json()["status"], "queued")
        job = self.client.get(
            f"/api/jobs/{first_start.json()['id']}",
            headers=operator_headers,
        )
        self.assertEqual(job.status_code, 200, job.text)
        self.assertEqual(job.json()["kind"], "compute")
        refreshed = self.client.get(
            f"/api/batches/{batch_id}", headers=operator_headers
        ).json()
        self.assertEqual(refreshed["jobs"]["compute"]["id"], job.json()["id"])
        self.assertEqual(refreshed["jobs"]["compute"]["status"], "queued")

    def test_batch_reads_bulk_load_exception_splits(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        batch_id = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "批量读取测试"},
        ).json()["id"]

        with self.app.state.database.session() as session:
            batch = session.get(Batch, batch_id)
            batch.status = "succeeded"
            source = BatchFile(
                batch_id=batch_id,
                original_name="批量读取测试.xlsx",
                storage_path="unused.xlsx",
                file_order=1,
                supplier_name="KuangBiao",
                supplier_code="GYS-023",
                delivery_total=10,
                import_total=0,
                manual_total=10,
                import_rows=[],
            )
            session.add(source)
            session.flush()
            for index in range(10):
                exception = ExceptionRecord(
                    batch_file_id=source.id,
                    sku="SKU-A",
                    original_site="US",
                    full_site="AMAZON:SEEKWAY:US",
                    destination="水鞋-广州仓",
                    delivery_quantity=1,
                    allocated_quantity=0,
                    manual_quantity=1,
                    reason=f"批量读取测试 {index}",
                    status="resolved",
                )
                session.add(exception)
                session.flush()
                session.add(
                    SplitRecord(
                        exception_id=exception.id,
                        quantity=1,
                        destination="水鞋-广州仓",
                        site="AMAZON:SEEKWAY:US",
                        supplier_code="GYS-023",
                        sku="SKU-A",
                        resolved=True,
                    )
                )
            session.commit()

        detail, detail_queries = self.get_with_query_count(
            f"/api/batches/{batch_id}",
            admin_headers,
        )
        listed, list_queries = self.get_with_query_count(
            "/api/batches",
            admin_headers,
        )
        original_reader = web_api_module.read_position_workbook
        with patch.object(
            web_api_module,
            "read_position_workbook",
            wraps=original_reader,
        ) as reader:
            exceptions, exception_queries = self.get_with_query_count(
                f"/api/batches/{batch_id}/exceptions",
                admin_headers,
            )
            repeated_exceptions = self.client.get(
                f"/api/batches/{batch_id}/exceptions",
                headers=admin_headers,
            )

        self.assertEqual(detail.status_code, 200, detail.text)
        self.assertEqual(listed.status_code, 200, listed.text)
        self.assertEqual(exceptions.status_code, 200, exceptions.text)
        self.assertEqual(
            repeated_exceptions.status_code,
            200,
            repeated_exceptions.text,
        )
        self.assertEqual(detail.json()["summary"]["import_total"], 10)
        listed_batch = next(
            batch for batch in listed.json() if batch["id"] == batch_id
        )
        self.assertEqual(
            listed_batch,
            {key: detail.json()[key] for key in listed_batch},
        )
        self.assertEqual(len(exceptions.json()), 10)
        self.assertEqual(repeated_exceptions.json(), exceptions.json())
        self.assertEqual(reader.call_count, 1)
        self.assertLessEqual(detail_queries, 15)
        self.assertLessEqual(list_queries, 8)
        self.assertLessEqual(exception_queries, 8)

    def test_batch_list_query_count_is_constant_as_batches_grow(self):
        admin_headers = self.login("admin", "admin-pass")
        version_ids = self.upload_active_versions(admin_headers)

        def add_batch(index: int) -> None:
            with self.app.state.database.session() as session:
                batch = Batch(
                    name=f"批次 {index}",
                    status="succeeded",
                    created_by=1,
                    purchase_version_id=version_ids["purchase"],
                    product_version_id=version_ids["product"],
                    supplier_version_id=version_ids["supplier"],
                    position_version_id=version_ids["position"],
                    template_version_id=version_ids["template"],
                )
                session.add(batch)
                session.flush()
                source = BatchFile(
                    batch_id=batch.id,
                    original_name=f"批次 {index}.xlsx",
                    storage_path=f"unused-{index}.xlsx",
                    file_order=1,
                    delivery_total=5,
                    import_total=2,
                    manual_total=3,
                    import_rows=[],
                )
                session.add(source)
                session.flush()
                exception = ExceptionRecord(
                    batch_file_id=source.id,
                    sku="SKU-A",
                    delivery_quantity=3,
                    allocated_quantity=0,
                    manual_quantity=3,
                    reason="数量超出采购余额",
                    status="resolved",
                )
                session.add(exception)
                session.flush()
                session.add_all(
                    [
                        SplitRecord(
                            exception_id=exception.id,
                            quantity=2,
                            destination="水鞋-广州仓",
                            site="AMAZON:SEEKWAY:US",
                            supplier_code="GYS-023",
                            sku="SKU-A",
                            resolved=True,
                        ),
                        SplitRecord(
                            exception_id=exception.id,
                            quantity=1,
                            sku="SKU-A",
                            resolved=False,
                        ),
                    ]
                )
                session.commit()

        add_batch(1)
        single, single_queries = self.get_with_query_count(
            "/api/batches",
            admin_headers,
        )
        for index in range(2, 11):
            add_batch(index)
        multiple, multiple_queries = self.get_with_query_count(
            "/api/batches",
            admin_headers,
        )

        self.assertEqual(single.status_code, 200, single.text)
        self.assertEqual(multiple.status_code, 200, multiple.text)
        self.assertEqual(single_queries, multiple_queries)
        self.assertLessEqual(multiple_queries, 8)
        self.assertEqual(
            [batch["name"] for batch in multiple.json()],
            [f"批次 {index}" for index in range(10, 0, -1)],
        )
        for batch in multiple.json():
            self.assertEqual(
                batch["summary"],
                {
                    "delivery_total": 5,
                    "import_total": 4,
                    "manual_total": 1,
                    "conserved": True,
                },
            )

    def test_position_frame_cache_evicts_least_recent_version(self):
        cache = web_api_module._PositionFrameCache(max_entries=2)
        loaded_versions = []

        def read_frame(path):
            version_id = int(path.stem)
            loaded_versions.append(version_id)
            return pd.DataFrame({"version_id": [version_id]})

        with patch.object(
            web_api_module,
            "read_position_workbook",
            side_effect=read_frame,
        ):
            first_frame = cache.get(1, Path("1.xlsx"))
            cache.get(2, Path("2.xlsx"))
            self.assertIs(cache.get(1, Path("1.xlsx")), first_frame)
            cache.get(3, Path("3.xlsx"))
            cache.get(2, Path("2.xlsx"))

        self.assertEqual(loaded_versions, [1, 2, 3, 2])

    def test_position_frame_cache_serializes_concurrent_misses(self):
        cache = web_api_module._PositionFrameCache(max_entries=2)
        concurrent_reads = Barrier(2)

        def read_frame(_path):
            try:
                concurrent_reads.wait(timeout=0.2)
            except BrokenBarrierError:
                pass
            return pd.DataFrame({"version_id": [1]})

        with patch.object(
            web_api_module,
            "read_position_workbook",
            side_effect=read_frame,
        ) as reader:
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = [
                    executor.submit(cache.get, 1, Path("1.xlsx")) for _ in range(2)
                ]
                frames = [future.result(timeout=5) for future in futures]

        self.assertEqual(reader.call_count, 1)
        self.assertIs(frames[0], frames[1])

    def test_concurrent_batch_uploads_get_distinct_contiguous_orders(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        batch_id = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "并发上传排序测试"},
        ).json()["id"]
        saved_uploads = Barrier(2)
        original_save_upload = web_api_module._save_upload

        async def synchronized_save_upload(*args, **kwargs):
            await original_save_upload(*args, **kwargs)
            saved_uploads.wait(timeout=5)

        def upload(filename: str):
            return self.client.post(
                f"/api/batches/{batch_id}/files",
                headers=admin_headers,
                files={"file": (filename, BytesIO(self.delivery_bytes()))},
            )

        filenames = ("KuangBiao-A.xlsx", "KuangBiao-B.xlsx")
        with (
            patch.object(
                web_api_module,
                "_save_upload",
                new=synchronized_save_upload,
            ),
            ThreadPoolExecutor(max_workers=2) as executor,
        ):
            responses = list(executor.map(upload, filenames))

        self.assertEqual(
            [response.status_code for response in responses],
            [201, 201],
            [response.text for response in responses],
        )
        batch = self.client.get(
            f"/api/batches/{batch_id}", headers=admin_headers
        ).json()
        self.assertEqual(
            [item["file_order"] for item in batch["files"]],
            [1, 2],
        )
        self.assertEqual(
            {item["original_name"] for item in batch["files"]},
            set(filenames),
        )

    def test_batch_file_limit_rejects_append_before_writing(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        batch_id = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "普通追加数量上限"},
        ).json()["id"]
        self.app.state.max_batch_upload_files = 1
        first = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=admin_headers,
            files={"file": ("first.xlsx", BytesIO(self.delivery_bytes()))},
        )
        self.assertEqual(first.status_code, 201, first.text)
        input_root = self.app.state.storage_root / "batches" / str(batch_id) / "inputs"
        files_before = set(input_root.iterdir())

        with patch.object(
            web_api_module,
            "_save_upload",
            wraps=web_api_module._save_upload,
        ) as save_upload:
            rejected = self.client.post(
                f"/api/batches/{batch_id}/files",
                headers=admin_headers,
                files={"file": ("second.xlsx", BytesIO(self.delivery_bytes()))},
            )

        self.assertEqual(rejected.status_code, 413, rejected.text)
        self.assertIn("最多上传 1 份", rejected.json()["detail"])
        save_upload.assert_not_awaited()
        self.assertEqual(set(input_root.iterdir()), files_before)
        with self.app.state.database.session() as session:
            sources = session.scalars(
                select(BatchFile).where(BatchFile.batch_id == batch_id)
            ).all()
        self.assertEqual([source.original_name for source in sources], ["first.xlsx"])

    def test_concurrent_batch_appends_enforce_file_limit_after_writing(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        batch_id = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "并发追加数量上限"},
        ).json()["id"]
        self.app.state.max_batch_upload_files = 1
        saved_uploads = Barrier(2)
        original_save_upload = web_api_module._save_upload

        async def synchronized_save_upload(*args, **kwargs):
            await original_save_upload(*args, **kwargs)
            saved_uploads.wait(timeout=5)

        def upload(filename: str):
            return self.client.post(
                f"/api/batches/{batch_id}/files",
                headers=admin_headers,
                files={"file": (filename, BytesIO(self.delivery_bytes()))},
            )

        with (
            patch.object(
                web_api_module,
                "_save_upload",
                new=synchronized_save_upload,
            ),
            ThreadPoolExecutor(max_workers=2) as executor,
        ):
            responses = list(executor.map(upload, ("first.xlsx", "second.xlsx")))

        self.assertEqual(
            sorted(response.status_code for response in responses),
            [201, 413],
            [response.text for response in responses],
        )
        rejected = next(
            response for response in responses if response.status_code == 413
        )
        self.assertIn("最多上传 1 份", rejected.json()["detail"])
        input_root = self.app.state.storage_root / "batches" / str(batch_id) / "inputs"
        self.assertEqual(len(list(input_root.iterdir())), 1)
        with self.app.state.database.session() as session:
            sources = session.scalars(
                select(BatchFile).where(BatchFile.batch_id == batch_id)
            ).all()
        self.assertEqual(len(sources), 1)

    def test_delivery_file_can_be_deleted_before_compute(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        batch_id = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "文件纠错测试"},
        ).json()["id"]
        first = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=admin_headers,
            files={
                "file": ("260717-狂飙-A交货单.xlsx", BytesIO(self.delivery_bytes()))
            },
        ).json()
        second = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=admin_headers,
            files={
                "file": ("260717-狂飙-B交货单.xlsx", BytesIO(self.delivery_bytes()))
            },
        ).json()
        with self.app.state.database.session() as session:
            removed_path = Path(session.get(BatchFile, second["id"]).storage_path)
        self.assertTrue(removed_path.is_file())

        deleted = self.client.request(
            "DELETE",
            f"/api/batches/{batch_id}/files/{second['id']}",
            headers=admin_headers,
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)
        self.assertEqual(deleted.json()["status"], "draft")
        self.assertEqual(
            [(item["id"], item["file_order"]) for item in deleted.json()["files"]],
            [(first["id"], 1)],
        )
        self.assertFalse(removed_path.exists())

        self.client.post(f"/api/batches/{batch_id}/preflight", headers=admin_headers)
        self.client.post(f"/api/batches/{batch_id}/compute", headers=admin_headers)
        blocked = self.client.request(
            "DELETE",
            f"/api/batches/{batch_id}/files/{first['id']}",
            headers=admin_headers,
        )
        self.assertEqual(blocked.status_code, 409)

    def test_preflight_rejects_invalid_excel_content(self):
        admin_headers = self.login("admin", "admin-pass")
        self.upload_active_versions(admin_headers)
        created = self.client.post(
            "/api/batches",
            headers=admin_headers,
            json={"name": "无效预检"},
        )
        batch_id = created.json()["id"]
        uploaded = self.client.post(
            f"/api/batches/{batch_id}/files",
            headers=admin_headers,
            files={"file": ("260717-狂飙-A交货单-发货10箱.xlsx", BytesIO(b"invalid"))},
        )
        self.assertEqual(uploaded.status_code, 201)

        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=admin_headers,
        )
        self.assertEqual(preflight.status_code, 400)
        batch = self.client.get(
            f"/api/batches/{batch_id}", headers=admin_headers
        ).json()
        self.assertEqual(batch["status"], "draft")

    def test_split_review_is_quantity_safe_and_invalidates_export(self):
        admin_headers = self.login("admin", "admin-pass")
        self.create_operator(admin_headers)
        self.upload_active_versions(admin_headers)
        operator_headers = self.login("operator", "operator-pass")
        batch_id = self.client.post(
            "/api/batches",
            headers=operator_headers,
            json={"name": "拆分测试"},
        ).json()["id"]

        with self.app.state.database.session() as session:
            batch = session.get(Batch, batch_id)
            batch.status = "succeeded"
            source = BatchFile(
                batch_id=batch_id,
                original_name="260717-狂飙-A交货单-发货10箱.xlsx",
                storage_path="unused.xlsx",
                file_order=1,
                supplier_name="KuangBiao",
                supplier_code="GYS-023",
                delivery_total=40,
                import_total=0,
                manual_total=40,
                document_note="260717-狂飙-01-10箱",
                import_rows=[],
            )
            session.add(source)
            session.flush()
            exception = ExceptionRecord(
                batch_file_id=source.id,
                sku="SKU-A",
                original_site="US",
                full_site="AMAZON:SEEKWAY:US",
                destination="水鞋-广州仓",
                delivery_quantity=40,
                allocated_quantity=0,
                manual_quantity=40,
                reason="超出采购未交量",
                status="pending",
            )
            session.add(exception)
            session.commit()
            exception_id = exception.id

        invalid = self.client.put(
            f"/api/exceptions/{exception_id}/split",
            headers=operator_headers,
            json={"parts": [{"quantity": 39, "resolved": False}]},
        )
        self.assertEqual(invalid.status_code, 400)

        valid = self.client.put(
            f"/api/exceptions/{exception_id}/split",
            headers=operator_headers,
            json={
                "parts": [
                    {"quantity": 25, "destination": "仓A", "resolved": True},
                    {"quantity": 15, "destination": "仓B", "resolved": False},
                ]
            },
        )
        self.assertEqual(valid.status_code, 200, valid.text)
        self.assertEqual(valid.json()["status"], "partial")
        self.assertEqual(
            [part["quantity"] for part in valid.json()["parts"]],
            [25, 15],
        )
        batch_after_split = self.client.get(
            f"/api/batches/{batch_id}", headers=operator_headers
        ).json()
        summary = batch_after_split["summary"]
        self.assertEqual(
            (
                summary["delivery_total"],
                summary["import_total"],
                summary["manual_total"],
            ),
            (40, 25, 15),
        )
        self.assertEqual(
            (
                batch_after_split["files"][0]["import_total"],
                batch_after_split["files"][0]["manual_total"],
            ),
            (25, 15),
        )

        export = self.client.post(
            f"/api/batches/{batch_id}/export",
            headers=operator_headers,
        )
        repeated = self.client.post(
            f"/api/batches/{batch_id}/export",
            headers=operator_headers,
        )
        self.assertEqual(export.status_code, 202, export.text)
        self.assertEqual(export.json()["id"], repeated.json()["id"])
        blocked_split = self.client.put(
            f"/api/exceptions/{exception_id}/split",
            headers=operator_headers,
            json={"parts": [{"quantity": 40, "resolved": False}]},
        )
        self.assertEqual(blocked_split.status_code, 409)


if __name__ == "__main__":
    unittest.main()
