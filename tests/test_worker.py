from contextlib import redirect_stderr
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from threading import Event, Lock, Thread
import subprocess
import sys
from tempfile import TemporaryDirectory
import time
import unittest
from unittest.mock import MagicMock, patch
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
import pandas as pd
from sqlalchemy import select

from delivery_note.pipeline import IMPORT_COLUMNS
from delivery_note.self_operated_inbound import INBOUND_TEMPLATE_COLUMNS
from tests.asgi_client import SyncASGIClient
import delivery_note.worker as worker_module
from delivery_note.web.api import create_app
from delivery_note.web.database import Database
from delivery_note.web.models import (
    Batch,
    BatchFile,
    ExceptionRecord,
    InputVersion,
    Job,
    PurchaseSyncJob,
    SelfOperatedInboundSyncJob,
)

try:
    from delivery_note.worker import (
        _consolidate_import_rows,
        _fail_job,
        _fetch_purchase_order_details,
        _fetch_incremental_purchase_order_details,
        build_parser,
        recover_stale_jobs,
        run_once,
    )
except ImportError:
    _consolidate_import_rows = None
    _fail_job = None
    recover_stale_jobs = None
    run_once = None


class WorkerExportConsolidationTests(unittest.TestCase):
    def test_purchase_details_use_eight_workers_and_keep_order(self):
        class DetailClient:
            def __init__(self):
                self.active = 0
                self.max_active = 0
                self.lock = Lock()

            def purchase_order_detail(self, po_code):
                with self.lock:
                    self.active += 1
                    self.max_active = max(self.max_active, self.active)
                time.sleep(0.03)
                with self.lock:
                    self.active -= 1
                return {"poCode": po_code}

        client = DetailClient()
        orders = [{"code": f"PO-{index}"} for index in range(8)]
        progress = []

        details = _fetch_purchase_order_details(
            client,
            orders,
            lambda count, po_code: progress.append((count, po_code)),
        )

        self.assertEqual(client.max_active, 8)
        self.assertEqual(
            [detail[1]["poCode"] for detail in details],
            [order["code"] for order in orders],
        )
        self.assertEqual(
            [count for count, _po_code in progress],
            list(range(1, 9)),
        )

    def test_incremental_details_reuse_cache_and_fallback_on_mismatch(self):
        from datetime import timezone

        from delivery_note.purchase_detail_cache import (
            build_purchase_detail_cache,
        )

        orders = [
            {"code": f"PO-{index}", "updateTime": "2026-08-27 08:00:00"}
            for index in range(20)
        ]
        details = [(order, {"poCode": order["code"], "balance": 1}) for order in orders]
        cached = build_purchase_detail_cache("source", details)["orders"]
        last_full = datetime.now(timezone.utc)

        class StableClient:
            def __init__(self):
                self.calls = []

            def purchase_order_detail(self, po_code):
                self.calls.append(po_code)
                return {"poCode": po_code, "balance": 1}

        stable = StableClient()
        fetched, stats, _ = _fetch_incremental_purchase_order_details(
            stable,
            orders,
            cached,
            last_full,
            lambda _count, _code: None,
            last_full,
        )

        self.assertEqual(len(stable.calls), 5)
        self.assertEqual(stats["cache_hit_count"], 15)
        self.assertEqual(stats["sampled_order_count"], 5)
        self.assertFalse(stats["incremental_fallback"])
        self.assertEqual([detail[1]["balance"] for detail in fetched], [1] * 20)

        class MismatchClient:
            def __init__(self):
                self.calls = []
                self.changed = False
                self.lock = Lock()

            def purchase_order_detail(self, po_code):
                with self.lock:
                    self.calls.append(po_code)
                    if not self.changed:
                        self.changed = True
                        return {"poCode": po_code, "balance": 2}
                return {"poCode": po_code, "balance": 1}

        mismatch = MismatchClient()
        fetched, stats, _ = _fetch_incremental_purchase_order_details(
            mismatch,
            orders,
            cached,
            last_full,
            lambda _count, _code: None,
            last_full,
        )

        self.assertEqual(len(mismatch.calls), 25)
        self.assertTrue(stats["incremental_fallback"])
        self.assertEqual(stats["sample_mismatch_count"], 1)
        self.assertEqual([detail[1]["balance"] for detail in fetched], [1] * 20)

    def test_multiple_delivery_notes_are_preserved_in_stable_order(self):
        self.assertIsNotNone(_consolidate_import_rows, "导入行合并函数尚未实现")
        if _consolidate_import_rows is None:
            return

        rows = pd.DataFrame(
            [
                ["仓A", "GYS-001", "SKU-A", 10, "站点A", "单据A", "原因"],
                ["仓A", "GYS-001", "SKU-A", 5, "站点A", "单据A", "原因：5"],
                ["仓A", "GYS-001", "SKU-A", 3, "站点A", "单据A", "其他备注"],
                ["仓A", "GYS-001", "SKU-A", 2, "站点A", "单据A", "原因"],
            ],
            columns=IMPORT_COLUMNS,
        )

        result = _consolidate_import_rows(rows)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["*本次交货量"], 20)
        self.assertEqual(result.iloc[0]["交货备注"], "原因：5；其他备注")


class WorkerIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.assertIsNotNone(run_once, "Worker 尚未实现")
        self.assertIsNotNone(recover_stale_jobs, "任务恢复尚未实现")
        if run_once is None or recover_stale_jobs is None:
            self.skipTest("Worker 尚未实现")
        self.directory = TemporaryDirectory()
        self.root = Path(self.directory.name)
        self.database_url = f"sqlite+pysqlite:///{self.root / 'worker.db'}"
        self.storage_root = self.root / "storage"
        self.app = create_app(
            database_url=self.database_url,
            storage_root=self.storage_root,
            bootstrap_admin=("admin", "admin-pass"),
        )
        self.client = SyncASGIClient(self.app)
        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin-pass"},
        )
        self.headers = {"Authorization": f"Bearer {login.json()['token']}"}
        self.inputs = self.create_master_inputs(self.root)
        self.upload_versions()

    def tearDown(self):
        if hasattr(self, "client"):
            self.client.close()
        if hasattr(self, "app"):
            self.app.state.database.dispose()
        if hasattr(self, "directory"):
            self.directory.cleanup()

    @staticmethod
    def create_master_inputs(root: Path) -> dict[str, Path]:
        purchase = root / "purchase.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"])
        sheet.append(
            ["待交货", "KuangBiao", "SKU-A", "AMAZON:SEEKWAY:US", "水鞋-广州仓", 100]
        )
        workbook.save(purchase)

        product = root / "product.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["SKU", "店铺/站点", "品类A", "锁仓MKSU"])
        sheet.append(["SKU-A", "SEEKWAY:US", "水鞋", "锁"])
        workbook.save(product)

        supplier = root / "supplier.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["供应商编号", "供应商名称", "状态"])
        sheet.append(["GYS-023", "KuangBiao", "启用"])
        workbook.save(supplier)

        position = root / "position.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "MSKU_视图"
        sheet.append(
            ["店铺-站点", "积加SKU", "MSKU", "规模定位", "备货定位", "已下单可售天数"]
        )
        sheet.append(["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 90])
        workbook.save(position)

        template = root / "template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "模板提示"
        sheet.merge_cells("A1:G1")
        sheet.append(IMPORT_COLUMNS)
        sheet.append(["示例仓", "示例供应商", "示例SKU", 1, "示例站点", "", ""])
        workbook.save(template)

        inbound_template = root / "inbound_template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "批量入库"
        sheet.append(INBOUND_TEMPLATE_COLUMNS)
        sheet.append(["示例"] + [""] * (len(INBOUND_TEMPLATE_COLUMNS) - 1))
        workbook.save(inbound_template)
        return {
            "purchase": purchase,
            "product": product,
            "supplier": supplier,
            "position": position,
            "template": template,
            "inbound_template": inbound_template,
        }

    @staticmethod
    def create_delivery(path: Path, quantity: int) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "汇总"
        sheet.append([])
        sheet.append([])
        sheet.append([])
        sheet.append(["SKU", "US站", "总计"])
        sheet.append(["SKU-A", quantity, quantity])
        workbook.save(path)
        return path

    @staticmethod
    def create_self_operated_delivery(path: Path, quantity: int) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "明细"
        sheet.append([])
        sheet.append([])
        sheet.append([])
        sheet.append(["积加SKU", "实收数量", "站点", "交货单号"])
        sheet.append(["SKU-A", quantity, "US站", "LN2608179025"])
        workbook.save(path)
        return path

    @staticmethod
    def create_self_operated_inbound(path: Path) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        source_columns = [
            column for column in INBOUND_TEMPLATE_COLUMNS if column != "最大可收货"
        ]
        sheet.append(source_columns + ["供应商"])
        for site, inbound_number in (
            ("AMAZON:RIVMOUNT:US", "IN-R"),
            ("AMAZON:SEEKWAY:US", "IN-S"),
        ):
            values = {column: "" for column in INBOUND_TEMPLATE_COLUMNS}
            values.update(
                {
                    "入库单号": inbound_number,
                    "入库仓": "自营仓",
                    "SKU": "SKU-A",
                    "平台站点": site,
                    "关联采购单": f"PO-20260801-{site}",
                    "关联交货单/调拨单": "LN2608179025",
                    "应收货": 10,
                }
            )
            sheet.append([values[column] for column in source_columns] + ["KuangBiao"])
        workbook.save(path)
        return path

    def upload_versions(self):
        for kind, path in self.inputs.items():
            with path.open("rb") as upload:
                response = self.client.post(
                    f"/api/input-versions/{kind}",
                    headers=self.headers,
                    data={"name": f"{kind}-v1", "activate": "true"},
                    files={"file": (path.name, upload)},
                )
            self.assertEqual(response.status_code, 201, response.text)

    @patch.dict(
        "os.environ",
        {
            "GERPGO_API_BASE_URL": "https://example.test",
            "GERPGO_APP_ID": "app",
            "GERPGO_APP_KEY": "key",
        },
    )
    @patch("delivery_note.worker.GerpgoClient.from_config")
    def test_purchase_sync_creates_inactive_candidate(self, client_factory):
        api_client = client_factory.return_value
        api_client.list_purchase_orders.return_value = [
            {"code": "PO-1", "invoicesStatusName": "待交货"}
        ]
        api_client.purchase_order_detail.return_value = {
            "warehouseProcureItemVos": [
                {
                    "procureItemVos": [
                        {
                            "product": "SKU-A",
                            "supplierCode": "GYS-023",
                            "supplierName": "接口供应商",
                            "arrivalMarketName": "共享",
                            "deliveryWarehouseName": "水鞋-广州仓",
                            "balanceQuantity": 12,
                        }
                    ]
                }
            ]
        }

        started = self.client.post(
            "/api/purchase-sync",
            headers=self.headers,
        )
        self.assertEqual(started.status_code, 201, started.text)
        job_id = started.json()["id"]

        completed_id = run_once(self.database_url, self.storage_root)

        self.assertEqual(completed_id, job_id)
        database = Database(self.database_url)
        try:
            with database.session() as session:
                job = session.get(PurchaseSyncJob, job_id)
                self.assertEqual(job.status, "succeeded", job.error_message)
                self.assertIsNotNone(job.candidate_version_id)
                candidate = session.get(
                    InputVersion,
                    job.candidate_version_id,
                )
                self.assertFalse(candidate.active)
                rows = pd.read_excel(candidate.storage_path)
                self.assertEqual(rows.iloc[0]["未交量"], 12)
                self.assertEqual(rows.iloc[0]["供应商"], "接口供应商")
                self.assertEqual(rows.iloc[0]["平台站点"], "共享")
                self.assertEqual(job.issues[0]["code"], "shared_site")
        finally:
            database.dispose()

    @patch.dict(
        "os.environ",
        {
            "GERPGO_API_BASE_URL": "https://example.test",
            "GERPGO_APP_ID": "app",
            "GERPGO_APP_KEY": "key",
        },
    )
    @patch("delivery_note.worker.GerpgoClient.from_config")
    def test_self_operated_inbound_sync_creates_candidate(self, client_factory):
        client_factory.return_value.list_self_operated_inbound_orders.return_value = [
            {
                "orderNo": "IN-1",
                "orderType": "purchase",
                "orderStatus": "WAIT_INBOUND",
                "warehouseName": "自营仓",
                "purchaseCode": "PO-1",
                "releatedCode": "LN2608179025",
                "supplierName": "KuangBiao",
                "orderItemResultList": [
                    {
                        "sku": "SKU-A",
                        "marketName": "共享",
                        "arriveNum": 10,
                    }
                ],
            }
        ]
        started = self.client.post(
            "/api/self-operated-inbound-sync",
            headers=self.headers,
        )
        self.assertEqual(started.status_code, 201, started.text)
        job_id = started.json()["id"]

        completed_id = run_once(self.database_url, self.storage_root)

        self.assertEqual(completed_id, job_id)
        database = Database(self.database_url)
        try:
            with database.session() as session:
                job = session.get(SelfOperatedInboundSyncJob, job_id)
                self.assertEqual(job.status, "succeeded", job.error_message)
                candidate = session.get(InputVersion, job.candidate_version_id)
                self.assertFalse(candidate.active)
                rows = pd.read_excel(candidate.storage_path)
                self.assertEqual(rows.iloc[0]["入库单号"], "IN-1")
                self.assertEqual(rows.iloc[0]["平台站点"], "共享")
                self.assertEqual(job.issues[0]["code"], "shared_site")
        finally:
            database.dispose()

    @patch(
        "delivery_note.worker._claim_self_operated_inbound_sync_job",
        return_value=None,
    )
    @patch("delivery_note.worker._claim_purchase_sync_job")
    @patch("delivery_note.worker._claim_job")
    def test_inbound_worker_only_claims_inbound_queue(
        self,
        claim_batch,
        claim_purchase,
        claim_inbound,
    ):
        completed_id = run_once(
            self.database_url,
            self.storage_root,
            "inbound-sync",
        )

        self.assertIsNone(completed_id)
        claim_batch.assert_not_called()
        claim_purchase.assert_not_called()
        claim_inbound.assert_called_once()

    def test_worker_queue_argument_defaults_to_all(self):
        parser = build_parser()

        self.assertEqual(parser.parse_args([]).queue, "all")
        self.assertEqual(parser.parse_args([]).max_attempts, 3)
        self.assertEqual(
            parser.parse_args(["--queue", "purchase-sync"]).queue,
            "purchase-sync",
        )
        with redirect_stderr(StringIO()), self.assertRaises(SystemExit):
            parser.parse_args(["--max-attempts", "0"])

    def create_batch(self, delivery_paths: list[Path]) -> tuple[int, int]:
        created = self.client.post(
            "/api/batches",
            headers=self.headers,
            json={"name": "Worker 集成测试"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        batch_id = created.json()["id"]
        for path in delivery_paths:
            with path.open("rb") as upload:
                response = self.client.post(
                    f"/api/batches/{batch_id}/files",
                    headers=self.headers,
                    files={"file": (path.name, upload)},
                )
            self.assertEqual(response.status_code, 201, response.text)
        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=self.headers,
        )
        self.assertEqual(preflight.status_code, 200, preflight.text)
        job = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=self.headers,
        )
        self.assertEqual(job.status_code, 202, job.text)
        return batch_id, job.json()["id"]

    def test_compute_split_export_and_zip_are_end_to_end(self):
        first = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx", 80
        )
        second = self.create_delivery(
            self.root / "260717-狂飙-B交货单-发货20箱.xlsx", 80
        )
        batch_id, compute_job_id = self.create_batch([first, second])

        self.assertEqual(run_once(self.database_url, self.storage_root), compute_job_id)
        batch = self.client.get(f"/api/batches/{batch_id}", headers=self.headers).json()
        self.assertEqual(batch["status"], "succeeded")
        illegal_preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight", headers=self.headers
        )
        self.assertEqual(illegal_preflight.status_code, 409)
        self.assertEqual(
            self.client.get(f"/api/batches/{batch_id}", headers=self.headers).json()[
                "status"
            ],
            "succeeded",
        )
        self.assertEqual(
            [(item["import_total"], item["manual_total"]) for item in batch["files"]],
            [(80, 0), (20, 60)],
        )
        exceptions = self.client.get(
            f"/api/batches/{batch_id}/exceptions", headers=self.headers
        ).json()
        self.assertEqual(len(exceptions), 1)
        self.assertEqual(exceptions[0]["manual_quantity"], 60)
        self.assertEqual(exceptions[0]["scale_position"], "短尾")
        self.assertEqual(exceptions[0]["stocking_position"], "备货")
        self.assertEqual(exceptions[0]["ordered_days"], 90)

        split = self.client.put(
            f"/api/exceptions/{exceptions[0]['id']}/split",
            headers=self.headers,
            json={
                "parts": [
                    {
                        "quantity": 25,
                        "destination": "水鞋-广州仓",
                        "delivery_note": "超出采购未交量",
                        "resolved": True,
                    },
                    {
                        "quantity": 35,
                        "delivery_note": "超出采购未交量",
                        "resolved": False,
                    },
                ]
            },
        )
        self.assertEqual(split.status_code, 200, split.text)
        self.assertEqual(split.json()["scale_position"], "短尾")
        self.assertEqual(split.json()["stocking_position"], "备货")
        self.assertEqual(split.json()["ordered_days"], 90)
        export = self.client.post(
            f"/api/batches/{batch_id}/export", headers=self.headers
        )
        self.assertEqual(export.status_code, 202, export.text)
        export_job_id = export.json()["id"]
        self.assertEqual(run_once(self.database_url, self.storage_root), export_job_id)

        completed = self.client.get(
            f"/api/batches/{batch_id}", headers=self.headers
        ).json()
        self.assertTrue(completed["download_ready"])
        self.assertTrue(completed["merged_download_ready"])
        self.assertTrue(all(item["download_ready"] for item in completed["files"]))
        merged_response = self.client.get(
            f"/api/batches/{batch_id}/download-merged", headers=self.headers
        )
        self.assertEqual(merged_response.status_code, 200, merged_response.text)
        merged_book = load_workbook(
            BytesIO(merged_response.content),
            data_only=True,
        )
        merged_import_sheet = merged_book["交货导入"]
        merged_pending_sheet = merged_book["待处理导入"]
        self.assertEqual(merged_import_sheet.cell(3, 1).value, "示例仓")
        self.assertEqual(
            sum(
                merged_import_sheet.cell(row, 4).value or 0
                for row in range(4, merged_import_sheet.max_row + 1)
            ),
            125,
        )
        self.assertEqual(
            sum(
                merged_pending_sheet.cell(row, 4).value or 0
                for row in range(3, merged_pending_sheet.max_row + 1)
            ),
            35,
        )
        merged_import_notes = [
            merged_import_sheet.cell(row, 6).value
            for row in range(4, merged_import_sheet.max_row + 1)
        ]
        self.assertTrue(merged_import_notes[0].endswith("-01-10箱"))
        self.assertTrue(
            all(note.endswith("-02-20箱") for note in merged_import_notes[1:])
        )
        self.assertTrue(merged_pending_sheet.cell(3, 6).value.endswith("-02-20箱"))
        archive_response = self.client.get(
            f"/api/batches/{batch_id}/download", headers=self.headers
        )
        self.assertEqual(archive_response.status_code, 200, archive_response.text)
        with ZipFile(BytesIO(archive_response.content)) as archive:
            names = sorted(archive.namelist())
            self.assertEqual(len(names), 2)
            self.assertFalse(any("merged" in name for name in names))
            second_book = load_workbook(BytesIO(archive.read(names[1])), data_only=True)

        import_sheet = second_book["交货导入"]
        pending_sheet = second_book["待处理导入"]
        import_total = sum(
            import_sheet.cell(row, 4).value or 0
            for row in range(4, import_sheet.max_row + 1)
        )
        pending_total = sum(
            pending_sheet.cell(row, 4).value or 0
            for row in range(3, pending_sheet.max_row + 1)
        )
        self.assertEqual((import_total, pending_total), (45, 35))
        import_records = [
            [import_sheet.cell(row, column).value for column in range(1, 8)]
            for row in range(4, import_sheet.max_row + 1)
        ]
        self.assertEqual(len(import_records), 1)
        self.assertEqual(import_records[0][3], 45)
        self.assertEqual(import_records[0][6], "超出采购未交量：60")
        self.assertIsNone(run_once(self.database_url, self.storage_root))
        repeated = self.client.post(
            f"/api/batches/{batch_id}/export", headers=self.headers
        )
        self.assertEqual(repeated.json()["id"], export_job_id)
        self.assertEqual(repeated.json()["status"], "succeeded")

        with self.app.state.database.session() as session:
            stored_batch = session.get(Batch, batch_id)
            first_export_dir = Path(stored_batch.zip_path).parent
            merged_path = Path(stored_batch.zip_path).with_name(
                f"batch-{batch_id}-merged.xlsx"
            )
        merged_path.unlink()
        self.assertEqual(
            self.client.get(
                f"/api/batches/{batch_id}/download-merged",
                headers=self.headers,
            ).status_code,
            404,
        )
        regenerated = self.client.post(
            f"/api/batches/{batch_id}/export", headers=self.headers
        )
        self.assertEqual(regenerated.json()["id"], export_job_id)
        self.assertEqual(regenerated.json()["status"], "queued")
        self.assertEqual(run_once(self.database_url, self.storage_root), export_job_id)
        self.assertEqual(
            self.client.get(
                f"/api/batches/{batch_id}/download-merged",
                headers=self.headers,
            ).status_code,
            200,
        )
        with self.app.state.database.session() as session:
            stored_batch = session.get(Batch, batch_id)
            second_export_dir = Path(stored_batch.zip_path).parent
        self.assertNotEqual(second_export_dir, first_export_dir)
        self.assertFalse(first_export_dir.exists())
        self.assertTrue(second_export_dir.is_dir())

        second_merged_path = second_export_dir / f"batch-{batch_id}-merged.xlsx"
        second_merged_path.unlink()
        retried = self.client.post(
            f"/api/batches/{batch_id}/export", headers=self.headers
        )
        self.assertEqual(retried.json()["status"], "queued")
        with (
            patch.object(
                worker_module.shutil,
                "rmtree",
                side_effect=PermissionError("denied"),
            ),
            self.assertLogs("delivery_note.worker", level="WARNING"),
        ):
            self.assertEqual(
                run_once(self.database_url, self.storage_root),
                export_job_id,
            )
        with self.app.state.database.session() as session:
            stored_batch = session.get(Batch, batch_id)
            stored_job = session.get(Job, export_job_id)
            third_export_dir = Path(stored_batch.zip_path).parent
            self.assertEqual(stored_batch.status, "succeeded")
            self.assertEqual(stored_job.status, "succeeded")
        self.assertNotEqual(third_export_dir, second_export_dir)
        self.assertTrue(third_export_dir.is_dir())
        self.assertTrue(second_export_dir.is_dir())

    def test_compute_uses_the_overreceipt_rule_locked_when_batch_was_created(self):
        first_rule = self.client.post(
            "/api/overreceipt-rule-versions",
            headers=self.headers,
            json={
                "name": "允许短尾超收 50",
                "short_tail_limit": 50,
                "medium_tail_limit": 20,
                "long_tail_limit": 10,
                "allowed_warehouses": ["水鞋-广州仓"],
            },
        )
        self.assertEqual(first_rule.status_code, 201, first_rule.text)
        first = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx", 80
        )
        second = self.create_delivery(
            self.root / "260717-狂飙-B交货单-发货20箱.xlsx", 80
        )
        batch_id, compute_job_id = self.create_batch([first, second])

        replacement_rule = self.client.post(
            "/api/overreceipt-rule-versions",
            headers=self.headers,
            json={
                "name": "停止自动超收",
                "short_tail_limit": 0,
                "medium_tail_limit": 0,
                "long_tail_limit": 0,
                "allowed_warehouses": [],
            },
        )
        self.assertEqual(replacement_rule.status_code, 201, replacement_rule.text)

        self.assertEqual(run_once(self.database_url, self.storage_root), compute_job_id)
        batch = self.client.get(f"/api/batches/{batch_id}", headers=self.headers).json()

        self.assertEqual(batch["overreceipt_rule"]["id"], first_rule.json()["id"])
        self.assertEqual(
            [(item["import_total"], item["manual_total"]) for item in batch["files"]],
            [(80, 0), (70, 10)],
        )
        self.assertEqual(
            batch["files"][1]["import_total"] + batch["files"][1]["manual_total"],
            80,
        )
        exceptions = self.client.get(
            f"/api/batches/{batch_id}/exceptions", headers=self.headers
        ).json()
        self.assertEqual(exceptions[0]["reason"], "超出允许超收量")
        self.assertEqual(exceptions[0]["manual_quantity"], 10)
        self.assertEqual(exceptions[0]["purchase_allocated_quantity"], 20)
        self.assertEqual(exceptions[0]["overreceipt_allocated_quantity"], 50)
        self.assertEqual(exceptions[0]["overreceipt_remaining_quantity"], 0)

        export = self.client.post(
            f"/api/batches/{batch_id}/export", headers=self.headers
        )
        self.assertEqual(export.status_code, 202, export.text)
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            export.json()["id"],
        )
        archive_response = self.client.get(
            f"/api/batches/{batch_id}/download", headers=self.headers
        )
        with ZipFile(BytesIO(archive_response.content)) as archive:
            names = sorted(archive.namelist())
            second_book = load_workbook(
                BytesIO(archive.read(names[1])),
                data_only=True,
            )
        import_sheet = second_book["交货导入"]
        pending_sheet = second_book["待处理导入"]
        self.assertEqual(
            [import_sheet.cell(2, column).value for column in range(1, 8)],
            IMPORT_COLUMNS,
        )
        self.assertEqual(
            sum(
                import_sheet.cell(row, 4).value or 0
                for row in range(4, import_sheet.max_row + 1)
            ),
            70,
        )
        self.assertEqual(pending_sheet.cell(3, 4).value, 10)
        self.assertEqual(pending_sheet.cell(3, 7).value, "超出允许超收量：10")

    def test_self_operated_site_choice_recomputes_and_exports_inbound_workbook(self):
        product = self.root / "product-ambiguous.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["SKU", "店铺/站点", "品类A", "锁仓MKSU"])
        sheet.append(["SKU-A", "RIVMOUNT:US", "水鞋", "锁"])
        sheet.append(["SKU-A", "SEEKWAY:US", "水鞋", "锁"])
        workbook.save(product)
        with product.open("rb") as upload:
            product_version = self.client.post(
                "/api/input-versions/product",
                headers=self.headers,
                data={"name": "product-ambiguous", "activate": "true"},
                files={"file": (product.name, upload)},
            )
        self.assertEqual(product_version.status_code, 201, product_version.text)
        rule = self.client.post(
            "/api/self-operated-overreceipt-rule-versions",
            headers=self.headers,
            json={"name": "自营仓超收 5 件", "allowance": 5},
        )
        self.assertEqual(rule.status_code, 201, rule.text)

        created = self.client.post(
            "/api/self-operated-batches",
            headers=self.headers,
            json={"name": "自营仓 Worker 集成测试"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        batch_id = created.json()["id"]
        delivery = self.create_self_operated_delivery(
            self.root / "260817-狂飙-质检交货单.xlsx",
            18,
        )
        inbound = self.create_self_operated_inbound(self.root / "自营仓收货入库单.xlsx")
        with delivery.open("rb") as upload:
            source = self.client.post(
                f"/api/batches/{batch_id}/files",
                headers=self.headers,
                files={"file": (delivery.name, upload)},
            )
        self.assertEqual(source.status_code, 201, source.text)
        with inbound.open("rb") as upload:
            inbound_upload = self.client.post(
                f"/api/self-operated-batches/{batch_id}/inbound-file",
                headers=self.headers,
                files={"file": (inbound.name, upload)},
            )
        self.assertEqual(inbound_upload.status_code, 200, inbound_upload.text)
        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=self.headers,
        )
        self.assertEqual(preflight.status_code, 200, preflight.text)
        compute = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=self.headers,
        )
        self.assertEqual(compute.status_code, 202, compute.text)
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            compute.json()["id"],
        )

        exceptions = self.client.get(
            f"/api/batches/{batch_id}/exceptions",
            headers=self.headers,
        ).json()
        self.assertEqual(len(exceptions), 1)
        self.assertEqual(exceptions[0]["reason"], "产品信息站点不唯一")
        self.assertIn("AMAZON:RIVMOUNT:US", exceptions[0]["full_site"])
        self.assertIn("AMAZON:SEEKWAY:US", exceptions[0]["full_site"])
        selected = self.client.put(
            f"/api/exceptions/{exceptions[0]['id']}/self-operated-site",
            headers=self.headers,
            json={"full_site": "AMAZON:RIVMOUNT:US"},
        )
        self.assertEqual(selected.status_code, 202, selected.text)
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            selected.json()["id"],
        )

        batch = self.client.get(
            f"/api/batches/{batch_id}",
            headers=self.headers,
        ).json()
        self.assertEqual(batch["summary"]["delivery_total"], 18)
        self.assertEqual(batch["summary"]["import_total"], 15)
        self.assertEqual(batch["summary"]["manual_total"], 3)
        exceptions = self.client.get(
            f"/api/batches/{batch_id}/exceptions",
            headers=self.headers,
        ).json()
        self.assertEqual(len(exceptions), 1)
        self.assertEqual(exceptions[0]["reason"], "超出允许超收量")
        self.assertEqual(exceptions[0]["manual_quantity"], 3)

        with self.app.state.database.session() as session:
            stored_batch = session.get(Batch, batch_id)
            stored_batch.error_message = "历史导出失败"
            session.commit()
        export = self.client.post(
            f"/api/batches/{batch_id}/export",
            headers=self.headers,
        )
        self.assertEqual(export.status_code, 202, export.text)
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            export.json()["id"],
        )
        refreshed_batch = self.client.get(
            f"/api/batches/{batch_id}", headers=self.headers
        ).json()
        self.assertIsNone(refreshed_batch["error_message"])
        result = self.client.get(
            f"/api/batch-files/{batch['files'][0]['id']}/download",
            headers=self.headers,
        )
        self.assertEqual(result.status_code, 200, result.text)
        output = load_workbook(BytesIO(result.content), data_only=True)
        output_sheet = output["批量入库"]
        self.assertEqual(output_sheet.cell(2, 8).value, "AMAZON:RIVMOUNT:US")
        self.assertEqual(output_sheet.cell(2, 12).value, 10)
        self.assertEqual(output_sheet.cell(2, 17).value, 15)
        self.assertEqual(output_sheet.cell(2, 18).value, "未分配库位")
        self.assertIsNone(output_sheet.cell(2, 19).value)
        self.assertEqual(output_sheet.cell(2, 20).value, "规则允许超收：5")

    def test_self_operated_multi_file_order_shares_balances_and_exports_all_formats(
        self,
    ):
        rule = self.client.post(
            "/api/self-operated-overreceipt-rule-versions",
            headers=self.headers,
            json={"name": "自营仓批次共享超收 5 件", "allowance": 5},
        )
        self.assertEqual(rule.status_code, 201, rule.text)
        created = self.client.post(
            "/api/self-operated-batches",
            headers=self.headers,
            json={"name": "自营仓多质检单 Worker 测试"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        batch_id = created.json()["id"]

        first_path = self.create_self_operated_delivery(
            self.root / "260817-狂飙-A质检交货单.xlsx",
            10,
        )
        second_path = self.create_self_operated_delivery(
            self.root / "260817-狂飙-B质检交货单.xlsx",
            8,
        )
        uploaded = []
        for source_path in (first_path, second_path):
            with source_path.open("rb") as upload:
                response = self.client.post(
                    f"/api/batches/{batch_id}/files",
                    headers=self.headers,
                    files={"file": (source_path.name, upload)},
                )
            self.assertEqual(response.status_code, 201, response.text)
            uploaded.append(response.json())

        reordered = self.client.put(
            f"/api/batches/{batch_id}/files/order",
            headers=self.headers,
            json={"file_ids": [uploaded[1]["id"], uploaded[0]["id"]]},
        )
        self.assertEqual(reordered.status_code, 200, reordered.text)
        inbound_path = self.create_self_operated_inbound(
            self.root / "自营仓多质检单待入库.xlsx"
        )
        with inbound_path.open("rb") as upload:
            inbound = self.client.post(
                f"/api/self-operated-batches/{batch_id}/inbound-file",
                headers=self.headers,
                files={"file": (inbound_path.name, upload)},
            )
        self.assertEqual(inbound.status_code, 200, inbound.text)

        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=self.headers,
        )
        self.assertEqual(preflight.status_code, 200, preflight.text)
        compute = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=self.headers,
        )
        self.assertEqual(compute.status_code, 202, compute.text)
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            compute.json()["id"],
        )

        computed = self.client.get(
            f"/api/batches/{batch_id}",
            headers=self.headers,
        ).json()
        self.assertEqual(
            [
                (
                    source["original_name"],
                    source["import_total"],
                    source["manual_total"],
                )
                for source in computed["files"]
            ],
            [
                ("260817-狂飙-B质检交货单.xlsx", 8, 0),
                ("260817-狂飙-A质检交货单.xlsx", 7, 3),
            ],
        )
        self.assertEqual(
            computed["summary"],
            {
                "delivery_total": 18,
                "import_total": 15,
                "manual_total": 3,
                "conserved": True,
            },
        )

        export = self.client.post(
            f"/api/batches/{batch_id}/export",
            headers=self.headers,
        )
        self.assertEqual(export.status_code, 202, export.text)
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            export.json()["id"],
        )
        exported = self.client.get(
            f"/api/batches/{batch_id}",
            headers=self.headers,
        ).json()
        self.assertTrue(exported["download_ready"])
        self.assertTrue(exported["merged_download_ready"])
        self.assertTrue(all(source["download_ready"] for source in exported["files"]))

        merged_response = self.client.get(
            f"/api/batches/{batch_id}/download-merged",
            headers=self.headers,
        )
        self.assertEqual(merged_response.status_code, 200, merged_response.text)
        merged_book = load_workbook(BytesIO(merged_response.content), data_only=True)
        merged_sheet = merged_book["批量入库"]
        self.assertEqual(merged_sheet.max_row, 2)
        self.assertEqual(
            sum(
                merged_sheet.cell(row, 17).value or 0
                for row in range(2, merged_sheet.max_row + 1)
            ),
            15,
        )

        archive_response = self.client.get(
            f"/api/batches/{batch_id}/download",
            headers=self.headers,
        )
        self.assertEqual(archive_response.status_code, 200, archive_response.text)
        with ZipFile(BytesIO(archive_response.content)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                [
                    "260817-狂飙-A质检交货单_积加入库.xlsx",
                    "260817-狂飙-B质检交货单_积加入库.xlsx",
                ],
            )

        for source in exported["files"]:
            response = self.client.get(
                f"/api/batch-files/{source['id']}/download",
                headers=self.headers,
            )
            self.assertEqual(response.status_code, 200, response.text)

        with self.app.state.database.session() as session:
            stored_batch = session.get(Batch, batch_id)
            first_export_dir = Path(stored_batch.zip_path).parent
            merged_path = first_export_dir / f"batch-{batch_id}-merged.xlsx"
        merged_path.unlink()
        regenerated = self.client.post(
            f"/api/batches/{batch_id}/export",
            headers=self.headers,
        )
        self.assertEqual(regenerated.status_code, 202, regenerated.text)
        self.assertEqual(regenerated.json()["status"], "queued")
        self.assertEqual(
            run_once(self.database_url, self.storage_root),
            export.json()["id"],
        )
        with self.app.state.database.session() as session:
            stored_batch = session.get(Batch, batch_id)
            second_export_dir = Path(stored_batch.zip_path).parent
        self.assertNotEqual(second_export_dir, first_export_dir)
        self.assertFalse(first_export_dir.exists())
        self.assertTrue(second_export_dir.is_dir())

    def test_previous_export_cleanup_has_strict_directory_boundary(self):
        export_root = self.storage_root / "batches" / "123" / "exports"
        current_dir = export_root / "export-current"
        old_dir = export_root / "export-old"
        referenced_dir = export_root / "export-referenced"
        non_export_dir = export_root / "keep-old"
        nested_dir = export_root / "nested" / "export-nested"
        outside_dir = self.storage_root / "export-outside"
        directories = [
            current_dir,
            old_dir,
            referenced_dir,
            non_export_dir,
            nested_dir,
            outside_dir,
        ]
        for directory in directories:
            directory.mkdir(parents=True)
            (directory / "artifact.xlsx").write_bytes(b"test")
        created = self.client.post(
            "/api/batches",
            headers=self.headers,
            json={"name": "旧导出目录引用保护"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        with self.app.state.database.session() as session:
            batch = session.get(Batch, created.json()["id"])
            batch.zip_path = str(referenced_dir / "artifact.xlsx")
            session.commit()

        worker_module._cleanup_previous_export_directories(
            self.app.state.database,
            export_root,
            current_dir,
            [directory / "artifact.xlsx" for directory in directories],
        )

        self.assertFalse(old_dir.exists())
        self.assertTrue(current_dir.is_dir())
        self.assertTrue(referenced_dir.is_dir())
        self.assertTrue(non_export_dir.is_dir())
        self.assertTrue(nested_dir.is_dir())
        self.assertTrue(outside_dir.is_dir())

    def test_self_operated_multi_file_compute_failure_persists_no_partial_result(self):
        created = self.client.post(
            "/api/self-operated-batches",
            headers=self.headers,
            json={"name": "自营仓多文件原子计算"},
        )
        self.assertEqual(created.status_code, 201, created.text)
        batch_id = created.json()["id"]
        source_paths = [
            self.create_self_operated_delivery(
                self.root / "260817-狂飙-原子-A.xlsx",
                5,
            ),
            self.create_self_operated_delivery(
                self.root / "260817-狂飙-原子-B.xlsx",
                5,
            ),
        ]
        for source_path in source_paths:
            with source_path.open("rb") as upload:
                response = self.client.post(
                    f"/api/batches/{batch_id}/files",
                    headers=self.headers,
                    files={"file": (source_path.name, upload)},
                )
            self.assertEqual(response.status_code, 201, response.text)
        inbound_path = self.create_self_operated_inbound(
            self.root / "自营仓原子计算待入库.xlsx"
        )
        with inbound_path.open("rb") as upload:
            response = self.client.post(
                f"/api/self-operated-batches/{batch_id}/inbound-file",
                headers=self.headers,
                files={"file": (inbound_path.name, upload)},
            )
        self.assertEqual(response.status_code, 200, response.text)
        preflight = self.client.post(
            f"/api/batches/{batch_id}/preflight",
            headers=self.headers,
        )
        self.assertEqual(preflight.status_code, 200, preflight.text)
        compute = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=self.headers,
        )
        self.assertEqual(compute.status_code, 202, compute.text)

        with self.app.state.database.session() as session:
            sources = session.scalars(
                select(BatchFile)
                .where(BatchFile.batch_id == batch_id)
                .order_by(BatchFile.file_order)
            ).all()
            Path(sources[1].storage_path).write_bytes(b"invalid")

        with self.assertLogs("delivery_note.worker", level="ERROR"):
            self.assertEqual(
                run_once(self.database_url, self.storage_root),
                compute.json()["id"],
            )
        with self.app.state.database.session() as session:
            batch = session.get(Batch, batch_id)
            sources = session.scalars(
                select(BatchFile).where(BatchFile.batch_id == batch_id)
            ).all()
            exceptions = session.scalars(
                select(ExceptionRecord)
                .join(BatchFile)
                .where(BatchFile.batch_id == batch_id)
            ).all()
            self.assertEqual(batch.status, "failed")
            self.assertEqual(
                [
                    (source.import_total, source.manual_total, source.import_rows)
                    for source in sources
                ],
                [(0, 0, []), (0, 0, [])],
            )
            self.assertEqual(exceptions, [])

    def test_stale_job_recovers_and_failed_batch_persists_no_partial_results(self):
        valid = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx", 80
        )
        second = self.create_delivery(
            self.root / "260717-狂飙-B交货单-发货20箱.xlsx", 20
        )
        batch_id, job_id = self.create_batch([valid, second])

        with self.app.state.database.session() as session:
            job = session.get(Job, job_id)
            job.status = "running"
            job.claim_token = "old-claim"
            job.heartbeat_at = datetime.utcnow() - timedelta(hours=2)
            sources = (
                session.query(BatchFile)
                .filter_by(batch_id=batch_id)
                .order_by(BatchFile.file_order)
                .all()
            )
            Path(sources[1].storage_path).write_bytes(b"not-an-excel-file")
            session.commit()
        recovered = recover_stale_jobs(
            self.database_url,
            stale_after=timedelta(minutes=30),
        )
        self.assertEqual(recovered, 1)
        with self.app.state.database.session() as session:
            self.assertIsNone(session.get(Job, job_id).claim_token)
        with self.assertLogs("delivery_note.worker", level="ERROR"):
            self.assertEqual(run_once(self.database_url, self.storage_root), job_id)

        with self.app.state.database.session() as session:
            batch = session.get(Batch, batch_id)
            sources = session.query(BatchFile).filter_by(batch_id=batch_id).all()
            exceptions = (
                session.query(ExceptionRecord)
                .join(BatchFile)
                .filter(BatchFile.batch_id == batch_id)
                .all()
            )
            job = session.get(Job, job_id)
            self.assertEqual(batch.status, "failed")
            self.assertEqual(job.status, "failed")
            self.assertTrue(job.error_message)
            self.assertEqual(
                [
                    (source.import_total, source.manual_total, source.import_rows)
                    for source in sources
                ],
                [(0, 0, []), (0, 0, [])],
            )
            self.assertEqual(exceptions, [])

    def test_stale_recovery_only_scans_the_selected_queue(self):
        delivery = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx",
            20,
        )
        _batch_id, batch_job_id = self.create_batch([delivery])
        stale_at = datetime.utcnow() - timedelta(hours=2)
        with self.app.state.database.session() as session:
            batch_job = session.get(Job, batch_job_id)
            batch_job.status = "running"
            batch_job.attempts = 1
            batch_job.claim_token = "batch-claim"
            batch_job.heartbeat_at = stale_at
            purchase_job = PurchaseSyncJob(
                status="running",
                active_slot=1,
                created_by=1,
                attempts=1,
                claim_token="purchase-claim",
                heartbeat_at=stale_at,
            )
            inbound_job = SelfOperatedInboundSyncJob(
                status="running",
                active_slot=1,
                created_by=1,
                attempts=1,
                claim_token="inbound-claim",
                heartbeat_at=stale_at,
            )
            session.add_all([purchase_job, inbound_job])
            session.commit()
            purchase_job_id = purchase_job.id
            inbound_job_id = inbound_job.id

        recovered = recover_stale_jobs(
            self.database_url,
            stale_after=timedelta(minutes=30),
            queue="purchase-sync",
        )

        self.assertEqual(recovered, 1)
        with self.app.state.database.session() as session:
            self.assertEqual(session.get(Job, batch_job_id).status, "running")
            self.assertEqual(
                session.get(PurchaseSyncJob, purchase_job_id).status,
                "queued",
            )
            self.assertEqual(
                session.get(SelfOperatedInboundSyncJob, inbound_job_id).status,
                "running",
            )

    def test_stale_jobs_stop_at_retry_cap_and_batch_can_be_retried_manually(self):
        delivery = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx",
            20,
        )
        batch_id, batch_job_id = self.create_batch([delivery])
        stale_at = datetime.utcnow() - timedelta(hours=2)
        with self.app.state.database.session() as session:
            batch = session.get(Batch, batch_id)
            batch.status = "running"
            batch_job = session.get(Job, batch_job_id)
            batch_job.status = "running"
            batch_job.attempts = 3
            batch_job.claim_token = "batch-claim"
            batch_job.heartbeat_at = stale_at
            purchase_job = PurchaseSyncJob(
                status="running",
                active_slot=1,
                created_by=1,
                attempts=3,
                claim_token="purchase-claim",
                heartbeat_at=stale_at,
            )
            inbound_job = SelfOperatedInboundSyncJob(
                status="running",
                active_slot=1,
                created_by=1,
                attempts=3,
                claim_token="inbound-claim",
                heartbeat_at=stale_at,
            )
            session.add_all([purchase_job, inbound_job])
            session.commit()
            purchase_job_id = purchase_job.id
            inbound_job_id = inbound_job.id

        recovered = recover_stale_jobs(
            self.database_url,
            stale_after=timedelta(minutes=30),
            max_attempts=3,
        )

        self.assertEqual(recovered, 3)
        with self.app.state.database.session() as session:
            batch = session.get(Batch, batch_id)
            batch_job = session.get(Job, batch_job_id)
            purchase_job = session.get(PurchaseSyncJob, purchase_job_id)
            inbound_job = session.get(
                SelfOperatedInboundSyncJob,
                inbound_job_id,
            )
            self.assertEqual(batch.status, "failed")
            self.assertEqual(batch_job.status, "failed")
            self.assertIsNotNone(batch_job.finished_at)
            self.assertEqual(purchase_job.status, "failed")
            self.assertIsNone(purchase_job.active_slot)
            self.assertEqual(inbound_job.status, "failed")
            self.assertIsNone(inbound_job.active_slot)

        retried = self.client.post(
            f"/api/batches/{batch_id}/compute",
            headers=self.headers,
        )
        self.assertEqual(retried.status_code, 202, retried.text)
        self.assertEqual(retried.json()["attempts"], 3)
        self.assertEqual(
            run_once(self.database_url, self.storage_root, queue="batch"),
            batch_job_id,
        )
        with self.app.state.database.session() as session:
            retried_job = session.get(Job, batch_job_id)
            self.assertEqual(retried_job.status, "succeeded")
            self.assertEqual(retried_job.attempts, 4)

    def test_background_heartbeat_runs_during_blocking_execution_and_stops(self):
        delivery = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx",
            20,
        )
        _batch_id, job_id = self.create_batch([delivery])
        execution_started = Event()
        release_execution = Event()
        heartbeat_seen = Event()
        original_heartbeat = worker_module._heartbeat

        def block_execution(*_args):
            execution_started.set()
            self.assertTrue(release_execution.wait(2))

        def observe_heartbeat(*args):
            original_heartbeat(*args)
            heartbeat_seen.set()

        result = []
        with (
            patch.object(worker_module, "LEASE_HEARTBEAT_INTERVAL_SECONDS", 0.01),
            patch.object(
                worker_module,
                "_execute_compute",
                side_effect=block_execution,
            ),
            patch.object(
                worker_module,
                "_heartbeat",
                side_effect=observe_heartbeat,
            ) as beat,
        ):
            thread = Thread(
                target=lambda: result.append(
                    run_once(self.database_url, self.storage_root, queue="batch")
                )
            )
            thread.start()
            self.assertTrue(execution_started.wait(2))
            self.assertTrue(heartbeat_seen.wait(2))
            release_execution.set()
            thread.join(timeout=2)
            self.assertFalse(thread.is_alive())
            heartbeat_count = beat.call_count
            time.sleep(0.05)
            self.assertEqual(beat.call_count, heartbeat_count)

        self.assertEqual(result, [job_id])

    def test_lost_lease_is_logged_and_cannot_be_overwritten(self):
        delivery = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx",
            20,
        )
        batch_id, job_id = self.create_batch([delivery])
        execution_started = Event()
        release_execution = Event()
        lease_lost = Event()
        original_heartbeat = worker_module._heartbeat
        original_process = worker_module.process_delivery_batch

        def block_execution(*args, **kwargs):
            execution_started.set()
            self.assertTrue(release_execution.wait(2))
            return original_process(*args, **kwargs)

        def observe_lease_loss(*args):
            try:
                original_heartbeat(*args)
            except worker_module.LostJobLeaseError:
                lease_lost.set()
                raise

        result = []
        with self.assertLogs("delivery_note.worker", level="ERROR") as logs:
            with (
                patch.object(
                    worker_module,
                    "LEASE_HEARTBEAT_INTERVAL_SECONDS",
                    0.01,
                ),
                patch.object(
                    worker_module,
                    "process_delivery_batch",
                    side_effect=block_execution,
                ),
                patch.object(
                    worker_module,
                    "_heartbeat",
                    side_effect=observe_lease_loss,
                ),
            ):
                thread = Thread(
                    target=lambda: result.append(
                        run_once(
                            self.database_url,
                            self.storage_root,
                            queue="batch",
                        )
                    )
                )
                thread.start()
                self.assertTrue(execution_started.wait(2))
                with self.app.state.database.session() as session:
                    job = session.get(Job, job_id)
                    original_claim_prefix = job.claim_token[:8]
                    job.claim_token = "replacement-claim"
                    batch = session.get(Batch, batch_id)
                    batch.status = "running"
                    session.commit()
                self.assertTrue(lease_lost.wait(2))
                release_execution.set()
                thread.join(timeout=2)
                self.assertFalse(thread.is_alive())

        self.assertEqual(result, [job_id])
        with self.app.state.database.session() as session:
            job = session.get(Job, job_id)
            batch = session.get(Batch, batch_id)
            self.assertEqual(job.status, "running")
            self.assertEqual(job.claim_token, "replacement-claim")
            self.assertEqual(batch.status, "running")
        log_output = "\n".join(logs.output)
        self.assertIn(f"queue=batch job_id={job_id}", log_output)
        self.assertIn(f"claim={original_claim_prefix}", log_output)
        self.assertIn("Traceback", log_output)

    def test_successful_finalize_stops_heartbeat_before_terminal_state(self):
        delivery = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx",
            20,
        )
        _batch_id, job_id = self.create_batch([delivery])
        heartbeat_after_terminal = Event()
        original_execute = worker_module._execute_compute
        original_heartbeat = worker_module._heartbeat

        def delay_after_finalize(*args):
            original_execute(*args)
            time.sleep(0.05)

        def observe_heartbeat(database, current_job_id, claim_token):
            with database.session() as session:
                if session.get(Job, current_job_id).status != "running":
                    heartbeat_after_terminal.set()
            original_heartbeat(database, current_job_id, claim_token)

        with (
            patch.object(worker_module, "LEASE_HEARTBEAT_INTERVAL_SECONDS", 0.005),
            patch.object(
                worker_module,
                "_execute_compute",
                side_effect=delay_after_finalize,
            ),
            patch.object(
                worker_module,
                "_heartbeat",
                side_effect=observe_heartbeat,
            ),
            patch.object(worker_module.LOGGER, "exception") as log_failure,
        ):
            completed_id = run_once(
                self.database_url,
                self.storage_root,
                queue="batch",
            )

        self.assertEqual(completed_id, job_id)
        self.assertFalse(heartbeat_after_terminal.is_set())
        log_failure.assert_not_called()
        with self.app.state.database.session() as session:
            self.assertEqual(session.get(Job, job_id).status, "succeeded")

    def test_stale_worker_cannot_overwrite_new_claim(self):
        delivery = self.create_delivery(
            self.root / "260717-狂飙-A交货单-发货10箱.xlsx", 20
        )
        batch_id, job_id = self.create_batch([delivery])
        with self.app.state.database.session() as session:
            job = session.get(Job, job_id)
            job.status = "running"
            job.claim_token = "new-claim"
            job.error_message = None
            batch = session.get(Batch, batch_id)
            batch.status = "running"
            session.commit()

        _fail_job(self.app.state.database, job_id, "old-claim", "旧 Worker 失败")

        with self.app.state.database.session() as session:
            job = session.get(Job, job_id)
            batch = session.get(Batch, batch_id)
            self.assertEqual(job.status, "running")
            self.assertEqual(job.claim_token, "new-claim")
            self.assertIsNone(job.error_message)
            self.assertEqual(batch.status, "running")


class WorkerProcessLifecycleTests(unittest.TestCase):
    def test_persistent_main_reuses_one_database_and_disposes_it(self):
        class StopAfterOnePoll:
            def __init__(self):
                self.wait_count = 0
                self.stopped = False

            def is_set(self):
                return self.stopped or self.wait_count > 0

            def wait(self, _timeout):
                self.wait_count += 1
                return False

            def set(self):
                self.stopped = True

        database = MagicMock()
        session = database.session.return_value.__enter__.return_value
        session.scalar.return_value = None
        session.scalars.return_value.all.return_value = []
        with (
            patch.object(worker_module, "Database", return_value=database) as factory,
            patch.object(worker_module, "Event", StopAfterOnePoll),
            patch.object(worker_module, "Thread") as thread_factory,
            patch.object(worker_module.signal, "signal", return_value=None),
        ):
            exit_code = worker_module.main(
                [
                    "--database-url",
                    "sqlite+pysqlite:///unused.db",
                    "--poll-interval",
                    "0.01",
                ]
            )

        self.assertEqual(exit_code, 0)
        factory.assert_called_once_with("sqlite+pysqlite:///unused.db")
        database.dispose.assert_called_once_with()
        thread_factory.return_value.start.assert_called_once_with()
        thread_factory.return_value.join.assert_called_once()

    def test_stale_watcher_continues_after_recovering_a_job(self):
        class StopAfterRecovery:
            def __init__(self):
                self.wait_count = 0

            def wait(self, _timeout):
                self.wait_count += 1
                return self.wait_count > 1

        database = MagicMock()
        with patch.object(
            worker_module,
            "_recover_stale_jobs",
            return_value=1,
            create=True,
        ) as recover:
            worker_module._watch_stale_jobs(
                database,
                timedelta(minutes=30),
                StopAfterRecovery(),
                "batch",
                3,
            )

        recover.assert_called_once_with(
            database,
            stale_after=timedelta(minutes=30),
            queue="batch",
            max_attempts=3,
        )

    def test_worker_exits_cleanly_on_sigterm(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            database_url = f"sqlite+pysqlite:///{root / 'worker.db'}"
            database = Database(database_url)
            database.create_schema()
            database.dispose()
            process = subprocess.Popen(
                [
                    sys.executable,
                    "-m",
                    "delivery_note.worker",
                    "--database-url",
                    database_url,
                    "--storage-root",
                    str(root / "storage"),
                    "--poll-interval",
                    "0.05",
                ],
                cwd=Path(__file__).resolve().parents[1],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            try:
                time.sleep(2)
                self.assertIsNone(process.poll(), "Worker 在收到信号前意外退出")
                process.terminate()
                returncode = process.wait(timeout=5)
                stdout, stderr = process.communicate()
            finally:
                if process.poll() is None:
                    process.kill()
                    process.wait(timeout=5)

        self.assertEqual(returncode, 0, f"stdout={stdout}\nstderr={stderr}")


if __name__ == "__main__":
    unittest.main()
