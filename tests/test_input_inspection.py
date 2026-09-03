from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from delivery_note.excel_io import (
    PRODUCT_COLUMNS,
    PURCHASE_COLUMNS,
    read_position_workbook,
)
from delivery_note.input_inspection import (
    inspect_input_version,
    inspect_input_version_with_preview,
    position_change_warnings,
    position_diff,
    preview_input_version,
    validate_position_frame,
    write_position_workbook,
)
from delivery_note.pipeline import IMPORT_COLUMNS, POSITION_SOURCE_COLUMNS


class InputInspectionTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.path = self.root / "position.xlsx"
        self.frame = pd.DataFrame(
            [["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 90]],
            columns=POSITION_SOURCE_COLUMNS,
        )

    def tearDown(self):
        self.temporary_directory.cleanup()

    def test_position_summary_preview_and_quality_issues(self):
        frame = pd.DataFrame(
            [
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 90],
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "未知", "", "many"],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )
        issues = validate_position_frame(frame)
        self.assertIn("duplicate_msku", {item["code"] for item in issues})
        self.assertIn("unknown_scale", {item["code"] for item in issues})
        self.assertIn("non_numeric_days", {item["code"] for item in issues})
        duplicate_issue = next(
            item for item in issues if item["code"] == "duplicate_msku"
        )
        self.assertEqual(duplicate_issue["row_numbers"], [2, 3])

        write_position_workbook(self.path, frame)
        inspection = inspect_input_version("position", self.path)
        self.assertEqual(inspection["row_count"], 2)
        self.assertEqual(inspection["metrics"], {"sites": 1, "skus": 1, "mskus": 1})
        self.assertEqual(
            {item["code"] for item in inspection["issues"]},
            {item["code"] for item in issues},
        )

        preview = preview_input_version("position", self.path, offset=1, limit=1)
        self.assertEqual(preview["total"], 2)
        self.assertEqual(preview["offset"], 1)
        self.assertEqual(preview["limit"], 1)
        self.assertEqual(preview["columns"], POSITION_SOURCE_COLUMNS)
        self.assertEqual(preview["rows"][0]["已下单可售天数"], "many")

    def test_combined_position_inspection_reads_the_workbook_once(self):
        write_position_workbook(self.path, self.frame)

        with patch(
            "delivery_note.input_inspection.read_position_workbook",
            wraps=read_position_workbook,
        ) as read_workbook:
            result = inspect_input_version_with_preview(
                "position",
                self.path,
                offset=0,
                limit=50,
            )

        self.assertEqual(read_workbook.call_count, 1)
        self.assertEqual(result["summary"]["row_count"], 1)
        self.assertEqual(result["preview"]["total"], 1)
        self.assertEqual(result["preview"]["rows"][0]["积加SKU"], "SKU-A")

    def test_combined_product_inspection_streams_only_required_columns(self):
        path = self.root / "product.xlsx"
        source_columns = [
            "其他字段",
            "锁仓MKSU",
            "SKU",
            "店铺/站点",
            "品类A",
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(source_columns)
        sheet.append(["忽略", "锁", "SKU-A", "SEEKWAY:US", "水鞋"])
        sheet.append(["忽略", "", "SKU-B", "SEEKWAY:CA", "配件"])
        sheet["A100"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        workbook.save(path)

        with patch(
            "delivery_note.input_inspection.read_product_workbook",
            side_effect=AssertionError("产品预览不应全量读取 DataFrame"),
        ):
            result = inspect_input_version_with_preview(
                "product",
                path,
                offset=1,
                limit=1,
            )

        expected_columns = [
            column for column in source_columns if column in PRODUCT_COLUMNS
        ]
        self.assertEqual(result["summary"]["row_count"], 2)
        self.assertEqual(result["summary"]["columns"], expected_columns)
        self.assertEqual(result["preview"]["total"], 2)
        self.assertEqual(result["preview"]["columns"], expected_columns)
        self.assertEqual(
            result["preview"]["rows"],
            [
                {
                    "锁仓MKSU": None,
                    "SKU": "SKU-B",
                    "店铺/站点": "SEEKWAY:CA",
                    "品类A": "配件",
                }
            ],
        )

    def test_purchase_inspection_marks_shared_site_as_warning(self):
        path = self.root / "purchase.xlsx"
        frame = pd.DataFrame(
            [
                ["待交货", "接口供应商", "SKU-A", "共享", "广州仓", 12],
                [
                    "交货中",
                    "接口供应商",
                    "SKU-B",
                    "AMAZON:SEEKWAY:US",
                    "广州仓",
                    8,
                ],
            ],
            columns=PURCHASE_COLUMNS,
        )
        frame.to_excel(path, index=False)

        inspection = inspect_input_version("purchase", path)
        combined = inspect_input_version_with_preview(
            "purchase",
            path,
            offset=0,
            limit=20,
        )

        issue = inspection["issues"][0]
        self.assertEqual(issue["severity"], "warning")
        self.assertEqual(issue["code"], "shared_site")
        self.assertEqual(issue["row_numbers"], [2])
        self.assertEqual(combined["summary"]["issues"], [issue])

    def test_written_position_workbook_round_trips(self):
        write_position_workbook(self.path, self.frame)
        self.assertEqual(
            read_position_workbook(self.path).to_dict("records"),
            self.frame.to_dict("records"),
        )

    def test_position_validation_reports_errors_and_warnings(self):
        frame = pd.DataFrame(
            [
                [None, "SKU-A", "MSKU-A", "短尾", "备货", 30],
                ["SEEKWAY:US", "", "MSKU-B", "中尾", None, 60],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )

        issues = {item["code"]: item for item in validate_position_frame(frame)}

        self.assertEqual(issues["empty_site"]["severity"], "error")
        self.assertEqual(issues["empty_sku"]["severity"], "error")
        self.assertEqual(issues["empty_stocking"]["severity"], "warning")
        self.assertEqual(issues["empty_site"]["row_numbers"], [2])
        self.assertEqual(issues["empty_sku"]["row_numbers"], [3])

    def test_position_diff_counts_composite_key_changes(self):
        base = pd.DataFrame(
            [
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 30],
                ["SEEKWAY:CA", "SKU-B", "MSKU-B", "中尾", "备货", 60],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )
        candidate = pd.DataFrame(
            [
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "不备货", 30],
                ["SEEKWAY:UK", "SKU-C", "MSKU-C", "长尾", "备货", 90],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )

        self.assertEqual(
            position_diff(base, candidate),
            {"added": 1, "modified": 1, "deleted": 1, "unchanged": 0},
        )

    def test_position_diff_normalizes_identity_keys(self):
        candidate = self.frame.copy()
        candidate.loc[0, "店铺-站点"] = " seekway:us "
        candidate.loc[0, "积加SKU"] = "sku-a"
        candidate.loc[0, "MSKU"] = "msku-a"

        self.assertEqual(
            position_diff(self.frame, candidate),
            {"added": 0, "modified": 0, "deleted": 0, "unchanged": 1},
        )

    def test_position_diff_stably_pairs_duplicate_identity_rows(self):
        base = pd.DataFrame(
            [
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 30],
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "中尾", "备货", 60],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )
        candidate = pd.DataFrame(
            [
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 30],
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "长尾", "备货", 90],
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "长尾", "备货", 90],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )

        self.assertEqual(
            position_diff(base, candidate),
            {"added": 1, "modified": 1, "deleted": 0, "unchanged": 1},
        )

    def test_position_change_warnings_cover_empty_and_fifty_percent_changes(self):
        base = pd.DataFrame(
            [
                [f"STORE:{index}", f"SKU-{index}", f"MSKU-{index}", "短尾", "备货", 30]
                for index in range(4)
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )
        empty = pd.DataFrame(columns=POSITION_SOURCE_COLUMNS)
        half = base.iloc[:2].copy()
        increased = pd.concat(
            [
                base,
                pd.DataFrame(
                    [
                        ["STORE:4", "SKU-4", "MSKU-4", "短尾", "备货", 30],
                        ["STORE:5", "SKU-5", "MSKU-5", "短尾", "备货", 30],
                    ],
                    columns=POSITION_SOURCE_COLUMNS,
                ),
            ],
            ignore_index=True,
        )

        self.assertEqual(
            {issue["code"] for issue in position_change_warnings(base, empty)},
            {"row_count_cleared", "sites_cleared", "skus_cleared"},
        )
        self.assertEqual(
            {issue["code"] for issue in position_change_warnings(base, half)},
            {"row_count_changed", "sites_changed", "skus_changed"},
        )
        self.assertEqual(
            {issue["code"] for issue in position_change_warnings(base, increased)},
            {"row_count_changed", "sites_changed", "skus_changed"},
        )

    def test_empty_msku_is_an_error_when_site_and_sku_have_multiple_rows(self):
        frame = pd.DataFrame(
            [
                ["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 30],
                [" seekway:us ", "sku-a", "", "中尾", "备货", 60],
            ],
            columns=POSITION_SOURCE_COLUMNS,
        )

        duplicate_issue = next(
            item
            for item in validate_position_frame(frame)
            if item["code"] == "duplicate_msku"
        )
        self.assertEqual(duplicate_issue["severity"], "error")
        self.assertEqual(duplicate_issue["row_numbers"], [3])

    def test_all_input_kinds_are_read_and_template_uses_second_row_headers(self):
        sources = {
            "purchase": pd.DataFrame(
                [["待交货", "KuangBiao", "SKU-A", "SEEKWAY:US", "广州仓", 10]],
                columns=["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"],
            ),
            "product": pd.DataFrame(
                [["SKU-A", "SEEKWAY:US", "水鞋", "锁"]],
                columns=["SKU", "店铺/站点", "品类A", "锁仓MKSU"],
            ),
            "supplier": pd.DataFrame(
                [["GYS-001", "KuangBiao", "启用"]],
                columns=["供应商编号", "供应商名称", "状态"],
            ),
        }
        for kind, frame in sources.items():
            with self.subTest(kind=kind):
                path = self.root / f"{kind}.xlsx"
                frame.to_excel(path, index=False)
                inspection = inspect_input_version(kind, path)
                self.assertEqual(inspection["row_count"], 1)
                self.assertEqual(inspection["columns"], list(frame.columns))
                preview = preview_input_version(kind, path, offset=0, limit=10)
                self.assertEqual(
                    preview["rows"][0][frame.columns[-1]], frame.iloc[0, -1]
                )

        template_path = self.root / "template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "模板提示"
        sheet.append(IMPORT_COLUMNS)
        sheet.append(["广州仓", "GYS-001", "SKU-A", 10, "SEEKWAY:US", None, None])
        workbook.save(template_path)

        inspection = inspect_input_version("template", template_path)
        preview = preview_input_version("template", template_path, offset=0, limit=10)
        self.assertEqual(inspection["columns"], IMPORT_COLUMNS)
        self.assertEqual(inspection["row_count"], 1)
        self.assertEqual(preview["rows"][0]["*本次交货量"], 10)
        self.assertIsNone(preview["rows"][0]["单据备注"])

    def test_unknown_input_kind_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "不支持的输入资料类型"):
            inspect_input_version("other", self.path)


if __name__ == "__main__":
    unittest.main()
