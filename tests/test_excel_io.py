from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

try:
    from delivery_note.excel_io import (
        read_delivery_workbook,
        read_supplier_workbook,
        write_delivery_workbook,
        write_exception_workbook,
        write_import_workbook,
    )
    from delivery_note.pipeline import (
        BatchResult,
        EXCEPTION_COLUMNS,
        IMPORT_COLUMNS,
        PENDING_COLUMNS,
    )
except ImportError:
    read_delivery_workbook = None
    read_supplier_workbook = None
    write_delivery_workbook = None
    write_exception_workbook = None
    write_import_workbook = None
    BatchResult = None
    EXCEPTION_COLUMNS = []
    IMPORT_COLUMNS = []
    PENDING_COLUMNS = []


class ExcelInputTests(unittest.TestCase):
    def test_read_delivery_workbook_current_layout(self):
        self.assertIsNotNone(read_delivery_workbook, "交货单读取函数尚未实现")
        if read_delivery_workbook is None:
            return

        with TemporaryDirectory() as directory:
            path = Path(directory) / "狂飙交货单.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "汇总"
            sheet.append([])
            sheet.append([])
            sheet.append([])
            sheet.append(["SKU", "CA站", "US站", "总计"])
            sheet.append(["SKU-A", 5, 10, 15])
            sheet.append(["总计", 5, 10, 15])
            workbook.save(path)

            result = read_delivery_workbook(path)

        self.assertEqual(
            result.to_dict("records"),
            [
                {"SKU": "SKU-A", "原始站点": "CA", "交货量": 5},
                {"SKU": "SKU-A", "原始站点": "US", "交货量": 10},
            ],
        )


    def test_read_added_supplier_xls(self):
        self.assertIsNotNone(read_supplier_workbook, "供应商资料读取函数尚未实现")
        if read_supplier_workbook is None:
            return

        source = Path(__file__).parent / "fixtures" / "supplier_minimal.xls"
        result = read_supplier_workbook(source)
        zhangdun = result[result["供应商名称"].eq("Zhangdun")].iloc[0]

        self.assertEqual(zhangdun["供应商编号"], "GYS-027")
        self.assertEqual(zhangdun["状态"], "启用")

class ExcelOutputTests(unittest.TestCase):
    def test_write_import_workbook_preserves_template_header_and_data_style(self):
        self.assertIsNotNone(write_import_workbook, "正式导入文件导出函数尚未实现")
        if write_import_workbook is None:
            return

        with TemporaryDirectory() as directory:
            directory_path = Path(directory)
            template_path = directory_path / "template.xlsx"
            output_path = directory_path / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "模板提示"
            sheet.merge_cells("A1:G1")
            sheet.append(IMPORT_COLUMNS)
            sheet.append(["示例仓", "示例供应商", "示例SKU", 1, "示例站点", "", ""])
            for cell in sheet[2]:
                cell.font = Font(color="FF0000", bold=True)
            for cell in sheet[3]:
                cell.font = Font(name="宋体", size=10, color="808080")
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            workbook.save(template_path)

            rows = pd.DataFrame(
                [
                    ["仓A", "KuangBiao", "SKU-A", 10, "AMAZON:SEEKWAY:US", "", ""],
                    ["仓B", "KuangBiao", "SKU-B", 20, "AMAZON:SEEKWAY:CA", "", ""],
                ],
                columns=IMPORT_COLUMNS,
            )
            write_import_workbook(template_path, output_path, rows)
            result = load_workbook(output_path)
            output_sheet = result.active

        self.assertEqual(output_sheet["A1"].value, "模板提示")
        self.assertEqual([cell.value for cell in output_sheet[2]], IMPORT_COLUMNS)
        self.assertEqual(output_sheet["A3"].value, "仓A")
        self.assertEqual(output_sheet["A4"].value, "仓B")
        self.assertEqual(output_sheet["D3"].value, 10)
        self.assertEqual(output_sheet["D4"].value, 20)
        self.assertEqual(output_sheet["A3"]._style, output_sheet["A4"]._style)
        self.assertNotEqual(output_sheet["A3"].value, "示例仓")

    def test_write_delivery_workbook_contains_import_details_and_editable_pending_rows(self):
        self.assertIsNotNone(write_delivery_workbook, "交货处理文件导出函数尚未实现")
        if write_delivery_workbook is None:
            return

        result = BatchResult(
            import_rows=pd.DataFrame(
                [["仓A", "KuangBiao", "SKU-A", 80, "AMAZON:SEEKWAY:US", "", ""]],
                columns=IMPORT_COLUMNS,
            ),
            exception_rows=pd.DataFrame(
                [
                    [
                        "SKU-A",
                        "US",
                        "AMAZON:SEEKWAY:US",
                        "水鞋-广州仓",
                        120,
                        80,
                        40,
                        "超出采购未交量",
                    ]
                ],
                columns=EXCEPTION_COLUMNS,
            ),
            delivery_total=120,
            import_total=80,
            manual_total=40,
        )

        with TemporaryDirectory() as directory:
            template_path = Path(directory) / "template.xlsx"
            output_path = Path(directory) / "delivery.xlsx"
            template_book = Workbook()
            template_sheet = template_book.active
            template_sheet["A1"] = "模板提示"
            template_sheet.merge_cells("A1:G1")
            template_sheet.append(IMPORT_COLUMNS)
            template_sheet.append(
                ["示例仓", "示例供应商", "示例SKU", 1, "示例站点", "", ""]
            )
            template_sheet.protection.sheet = True
            template_book.save(template_path)
            pending_rows = pd.DataFrame(
                [
                    [
                        "水鞋-广州仓",
                        "KuangBiao",
                        "SKU-A",
                        40,
                        "AMAZON:SEEKWAY:US",
                        "狂飙交货单",
                        "超出采购未交量：40",
                        '{"MSKU-B":"短尾","MSKU-A":"长尾"}',
                        '{"MSKU-B":"备货","MSKU-A":"不备货"}',
                        '{"MSKU-B":120,"MSKU-A":30}',
                    ]
                ],
                columns=PENDING_COLUMNS,
            )
            write_delivery_workbook(
                template_path,
                output_path,
                result,
                result.import_rows,
                pending_rows,
            )
            workbook = load_workbook(output_path, data_only=True)

        self.assertEqual(workbook.sheetnames, ["交货导入", "待处理导入"])
        self.assertNotIn("运行汇总", workbook.sheetnames)
        self.assertNotIn("异常明细", workbook.sheetnames)
        self.assertEqual(workbook["交货导入"]["A1"].value, "模板提示")
        self.assertEqual(workbook["交货导入"]["A3"].value, "仓A")
        self.assertTrue(workbook["交货导入"].protection.sheet)
        self.assertTrue(workbook["交货导入"]["A1"].protection.locked)
        self.assertTrue(workbook["交货导入"]["A2"].protection.locked)
        self.assertFalse(workbook["交货导入"]["A3"].protection.locked)
        self.assertFalse(workbook["交货导入"]["G3"].protection.locked)
        default_protection_id = workbook._cell_styles[0].protectionId
        self.assertFalse(workbook._protections[default_protection_id].locked)
        self.assertEqual(workbook["待处理导入"]["A1"].value, "模板提示")
        self.assertEqual(
            [cell.value for cell in workbook["待处理导入"][2]], PENDING_COLUMNS
        )
        self.assertEqual(workbook["待处理导入"]["A3"].value, "水鞋-广州仓")
        self.assertEqual(workbook["待处理导入"]["F3"].value, "狂飙交货单")
        self.assertTrue(workbook["待处理导入"].protection.sheet)
        self.assertTrue(workbook["待处理导入"]["A1"].protection.locked)
        self.assertTrue(workbook["待处理导入"]["A2"].protection.locked)
        self.assertFalse(workbook["待处理导入"]["A3"].protection.locked)
        self.assertFalse(workbook["待处理导入"]["G3"].protection.locked)

        self.assertEqual(
            workbook["待处理导入"]["H3"].value,
            '{"MSKU-B":"短尾","MSKU-A":"长尾"}',
        )
        self.assertEqual(
            workbook["待处理导入"]["J3"].value,
            '{"MSKU-B":120,"MSKU-A":30}',
        )
        self.assertTrue(workbook["待处理导入"]["H2"].protection.locked)
        self.assertFalse(workbook["待处理导入"]["H3"].protection.locked)
        self.assertTrue(workbook["待处理导入"]["H3"].alignment.wrap_text)
        self.assertEqual(workbook["待处理导入"].row_dimensions[3].height, 60)
        self.assertEqual(workbook["待处理导入"].column_dimensions["G"].width, 35)

if __name__ == "__main__":
    unittest.main()
