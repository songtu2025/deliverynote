from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection

from .pipeline import (
    BatchResult,
    EXCEPTION_COLUMNS,
    IMPORT_COLUMNS,
    PENDING_COLUMNS,
    POSITION_SOURCE_COLUMNS,
    POSITION_VALUE_COLUMNS,
    normalize_delivery_sheet,
)


PURCHASE_COLUMNS = ["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"]
PRODUCT_COLUMNS = ["SKU", "店铺/站点", "品类A", "锁仓MKSU"]
SUPPLIER_COLUMNS = ["供应商编号", "供应商名称", "状态"]


def read_delivery_workbook(path: Path) -> pd.DataFrame:
    sheet = pd.read_excel(path, sheet_name="汇总", skiprows=3)
    return normalize_delivery_sheet(sheet)


def read_product_workbook(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, usecols=PRODUCT_COLUMNS)


def read_purchase_workbook(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, usecols=PURCHASE_COLUMNS)


def read_supplier_workbook(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, usecols=SUPPLIER_COLUMNS)

def read_position_workbook(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name="MSKU_视图", usecols=POSITION_SOURCE_COLUMNS)


def validate_template_workbook(path: Path) -> None:
    """只读校验官方模板的 A:G 表头和示例格式行。"""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        headers = [sheet.cell(row=2, column=column).value for column in range(1, 8)]
        if headers != IMPORT_COLUMNS:
            raise ValueError("官方模板表头与预期字段不一致")
        if sheet.max_row < 3:
            raise ValueError("官方模板缺少第 3 行示例格式")
    finally:
        workbook.close()


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    return value.item() if hasattr(value, "item") else value


def _populate_import_sheet(
    sheet,
    import_rows: pd.DataFrame,
    *,
    preserve_example_row: bool = False,
) -> None:
    if list(import_rows.columns) != IMPORT_COLUMNS:
        raise ValueError("正式导入数据字段与官方模板不一致")

    template_headers = [sheet.cell(row=2, column=column).value for column in range(1, 8)]
    if template_headers != IMPORT_COLUMNS:
        raise ValueError("官方模板表头与预期字段不一致")
    if sheet.max_row < 3:
        raise ValueError("官方模板缺少第 3 行示例格式")

    styles = [copy(sheet.cell(row=3, column=column)._style) for column in range(1, 8)]
    row_height = sheet.row_dimensions[3].height
    if sheet.max_row > 3:
        sheet.delete_rows(4, sheet.max_row - 3)
    first_data_row = 4 if preserve_example_row else 3
    if not preserve_example_row:
        for column in range(1, 8):
            sheet.cell(row=3, column=column).value = None

    for row_offset, values in enumerate(import_rows.itertuples(index=False, name=None)):
        row_number = first_data_row + row_offset
        if row_number > 3:
            sheet.row_dimensions[row_number].height = row_height
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column)
            cell._style = copy(styles[column - 1])
            cell.value = _excel_value(value)
    for row_number in range(first_data_row, first_data_row + len(import_rows)):
        sheet.cell(row=row_number, column=4).number_format = "0"


def write_import_workbook(
    template_path: Path,
    output_path: Path,
    import_rows: pd.DataFrame,
) -> None:
    """复制官方模板，在第 3 行起写入可导入数据。"""
    workbook = load_workbook(template_path)
    _populate_import_sheet(workbook.active, import_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _style_table_header(cells) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in cells:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _fit_columns(sheet, maximum_width: int = 50) -> None:
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max((len(value) for value in values), default=8) + 2, maximum_width)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 10)


def write_exception_workbook(
    template_path: Path,
    output_path: Path,
    result: BatchResult,
    metadata: dict[str, Any],
    pending_rows: pd.DataFrame,
) -> None:
    """输出运行汇总、异常明细和可二次导入数据。"""
    workbook = load_workbook(template_path)
    pending_sheet = workbook.active
    pending_sheet.title = "待处理导入"
    _populate_import_sheet(pending_sheet, pending_rows)

    summary = workbook.create_sheet("运行汇总", 0)
    summary_rows = [
        *metadata.items(),
        ("交货总量", result.delivery_total),
        ("自动导入量", result.import_total),
        ("人工处理量", result.manual_total),
    ]
    for label, value in summary_rows:
        summary.append([label, _excel_value(value)])
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in summary["A"]:
        cell.fill = label_fill
        cell.font = Font(bold=True, color="1F1F1F")
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 48
    for row in range(1, summary.max_row + 1):
        summary.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)

    details = workbook.create_sheet("异常明细", 1)
    details.append(EXCEPTION_COLUMNS)
    _style_table_header(details[1])
    for values in result.exception_rows[EXCEPTION_COLUMNS].itertuples(
        index=False,
        name=None,
    ):
        details.append([_excel_value(value) for value in values])
    details.freeze_panes = "A2"
    details.auto_filter.ref = f"A1:H{max(details.max_row, 1)}"
    for column in (5, 6, 7):
        for row in range(2, details.max_row + 1):
            details.cell(row=row, column=column).number_format = "#,##0"
    _fit_columns(details)
    workbook.active = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)



def _set_default_cells_unlocked(workbook) -> None:
    """让未显式设置样式的空白单元格默认可编辑。"""
    unlocked_id = workbook._protections.add(Protection(locked=False))
    workbook._cell_styles[0].protectionId = unlocked_id
    workbook._cell_styles._rebuild_dict()


def _protect_header_only(sheet) -> None:
    """锁定前两行内容，同时允许用户调整导出文件格式。"""
    sheet.protection.sheet = True
    sheet.protection.formatCells = False
    sheet.protection.formatColumns = False
    sheet.protection.formatRows = False
    sheet.protection.selectLockedCells = False
    for row in sheet.iter_rows(
        min_row=1, max_row=2, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell._style = copy(cell._style)
            cell.protection = Protection(locked=True, hidden=True)
    for row in sheet.iter_rows(
        min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell._style = copy(cell._style)
            protection = copy(cell.protection)
            protection.locked = False
            cell.protection = protection



def _populate_pending_position_columns(sheet, pending_rows: pd.DataFrame) -> None:
    if list(pending_rows.columns) != PENDING_COLUMNS:
        raise ValueError("待处理导入数据字段与预期不一致")
    sheet.column_dimensions["G"].width = 35

    header_style = copy(sheet.cell(row=2, column=7)._style)
    data_style = copy(sheet.cell(row=3, column=7)._style)
    row_count = max(len(pending_rows), 1)
    for column, (letter, header) in enumerate(
        zip(("H", "I", "J"), POSITION_VALUE_COLUMNS), start=8
    ):
        header_cell = sheet.cell(row=2, column=column)
        header_cell._style = copy(header_style)
        header_cell.value = header
        header_cell.alignment = copy(header_cell.alignment)
        header_cell.alignment = Alignment(
            horizontal=header_cell.alignment.horizontal,
            vertical=header_cell.alignment.vertical,
            wrap_text=True,
        )
        sheet.column_dimensions[letter].width = 45
        for row_offset in range(row_count):
            cell = sheet.cell(row=3 + row_offset, column=column)
            cell._style = copy(data_style)
            if row_offset < len(pending_rows):
                cell.value = _excel_value(pending_rows.iloc[row_offset][header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header == "已下单可售天数" and not isinstance(cell.value, str):
                cell.number_format = "0.00"

    for row_offset in range(len(pending_rows)):
        helper_values = pending_rows.iloc[row_offset][POSITION_VALUE_COLUMNS]
        if any(
            isinstance(value, str) and value.startswith("{") for value in helper_values
        ):
            sheet.row_dimensions[3 + row_offset].height = 60

def write_delivery_workbook(
    template_path: Path,
    output_path: Path,
    result: BatchResult,
    import_rows: pd.DataFrame,
    pending_rows: pd.DataFrame,
) -> None:
    """输出交货导入、异常明细和可编辑的待处理导入数据。"""
    workbook = load_workbook(template_path)
    import_sheet = workbook.active
    import_sheet.title = "交货导入"
    pending_sheet = workbook.copy_worksheet(import_sheet)
    pending_sheet.title = "待处理导入"
    pending_sheet.protection = copy(import_sheet.protection)

    _set_default_cells_unlocked(workbook)
    _populate_import_sheet(
        import_sheet,
        import_rows,
        preserve_example_row=True,
    )
    _populate_import_sheet(pending_sheet, pending_rows[IMPORT_COLUMNS])
    _populate_pending_position_columns(pending_sheet, pending_rows)
    _protect_header_only(import_sheet)
    _protect_header_only(pending_sheet)
    workbook.active = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
