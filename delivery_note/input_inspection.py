from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import Workbook, load_workbook

from .excel_io import (
    PRODUCT_COLUMNS,
    PURCHASE_COLUMNS,
    read_position_workbook,
    read_product_workbook,
    read_purchase_workbook,
    read_supplier_workbook,
    validate_template_workbook,
)
from .pipeline import POSITION_SOURCE_COLUMNS


POSITION_KEY = ["店铺-站点", "积加SKU", "MSKU"]
_POSITION_VALUES = [column for column in POSITION_SOURCE_COLUMNS if column not in POSITION_KEY]
_KNOWN_SCALES = {"短尾", "中尾", "长尾"}
_STREAMING_COLUMNS = {
    "product": PRODUCT_COLUMNS,
    "purchase": PURCHASE_COLUMNS,
}


def _read_template_workbook(path: Path) -> pd.DataFrame:
    validate_template_workbook(path)
    return pd.read_excel(path, header=1, usecols="A:G")


def _read_frame(kind: str, path: Path) -> pd.DataFrame:
    readers: dict[str, Callable[[Path], pd.DataFrame]] = {
        "purchase": read_purchase_workbook,
        "product": read_product_workbook,
        "supplier": read_supplier_workbook,
        "position": read_position_workbook,
        "template": _read_template_workbook,
    }
    try:
        reader = readers[kind]
    except KeyError as error:
        raise ValueError(f"不支持的输入资料类型：{kind}") from error
    return reader(Path(path))


def _json_safe(value: Any) -> Any:
    if value is None or bool(pd.isna(value)):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "item"):
        return _json_safe(value.item())
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _stream_json_safe(value: Any) -> Any:
    if value == "":
        return None
    return _json_safe(value)


def _inspect_frame(kind: str, frame: pd.DataFrame) -> dict:
    result = {
        "kind": kind,
        "row_count": len(frame),
        "columns": [str(column) for column in frame.columns],
        "metrics": {},
        "issues": [],
    }
    if kind == "position":
        result["metrics"] = {
            "sites": int(frame["店铺-站点"].dropna().astype(str).str.strip().nunique()),
            "skus": int(frame["积加SKU"].dropna().astype(str).str.strip().nunique()),
            "mskus": int(frame["MSKU"].dropna().astype(str).str.strip().nunique()),
        }
        result["issues"] = validate_position_frame(frame)
    return result


def _preview_frame(
    kind: str,
    frame: pd.DataFrame,
    offset: int,
    limit: int,
) -> dict:
    page = frame.iloc[offset : offset + limit]
    columns = [str(column) for column in frame.columns]
    rows = [
        {
            str(column): _json_safe(value)
            for column, value in zip(frame.columns, values)
        }
        for values in page.itertuples(index=False, name=None)
    ]
    return {
        "kind": kind,
        "columns": columns,
        "rows": rows,
        "total": len(frame),
        "offset": offset,
        "limit": limit,
    }


def _stream_xlsx_inspection(
    kind: str,
    path: Path,
    offset: int,
    limit: int,
) -> dict:
    """流式读取预览行，同时计算有效数据总行数。"""

    expected_columns = _STREAMING_COLUMNS[kind]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        values = sheet.iter_rows(values_only=True)
        header = next(values, ())
        header_names = [
            "" if value is None else str(value)
            for value in header
        ]
        expected_set = set(expected_columns)
        selected_columns = [
            (index, name)
            for index, name in enumerate(header_names)
            if name in expected_set
        ]
        found = {name for _index, name in selected_columns}
        missing = [
            column for column in expected_columns if column not in found
        ]
        if missing:
            raise ValueError(f"缺少必要字段：{', '.join(missing)}")

        page_rows: list[tuple[int, dict[str, Any]]] = []
        last_data_offset = -1
        for row_offset, row in enumerate(values):
            selected_values = [
                row[index] if index < len(row) else None
                for index, _name in selected_columns
            ]
            if any(value not in (None, "") for value in selected_values):
                last_data_offset = row_offset
            if offset <= row_offset < offset + limit:
                page_rows.append(
                    (
                        row_offset,
                        {
                            name: _stream_json_safe(value)
                            for (_index, name), value in zip(
                                selected_columns,
                                selected_values,
                            )
                        },
                    )
                )

        total = last_data_offset + 1
        columns = [name for _index, name in selected_columns]
        preview = {
            "kind": kind,
            "columns": columns,
            "rows": [
                row
                for row_offset, row in page_rows
                if row_offset < total
            ],
            "total": total,
            "offset": offset,
            "limit": limit,
        }
        summary = {
            "kind": kind,
            "row_count": total,
            "columns": columns,
            "metrics": {},
            "issues": [],
        }
        return {"summary": summary, "preview": preview}
    finally:
        workbook.close()


def inspect_input_version(kind: str, path: Path) -> dict:
    return _inspect_frame(kind, _read_frame(kind, path))


def inspect_input_version_with_preview(
    kind: str,
    path: Path,
    offset: int,
    limit: int,
) -> dict:
    """一次读取基础资料并生成摘要与分页预览。"""

    path = Path(path)
    if (
        kind in _STREAMING_COLUMNS
        and path.suffix.lower() in {".xlsx", ".xlsm"}
    ):
        return _stream_xlsx_inspection(kind, path, offset, limit)
    frame = _read_frame(kind, path)
    return {
        "summary": _inspect_frame(kind, frame),
        "preview": _preview_frame(kind, frame, offset, limit),
    }


def preview_input_version(
    kind: str,
    path: Path,
    offset: int,
    limit: int,
) -> dict:
    return inspect_input_version_with_preview(
        kind,
        path,
        offset,
        limit,
    )["preview"]


def _text_values(frame: pd.DataFrame, column: str) -> pd.Series:
    return frame[column].fillna("").astype(str).str.strip()


def _identity_values(frame: pd.DataFrame, column: str) -> pd.Series:
    return _text_values(frame, column).str.upper()


def _row_numbers(mask: pd.Series) -> list[int]:
    return [position + 2 for position, selected in enumerate(mask) if bool(selected)]


def _append_issue(
    issues: list[dict],
    *,
    severity: str,
    code: str,
    message: str,
    mask: pd.Series,
) -> None:
    row_numbers = _row_numbers(mask)
    if row_numbers:
        issues.append(
            {
                "severity": severity,
                "code": code,
                "message": message,
                "row_numbers": row_numbers,
            }
        )


def validate_position_frame(frame: pd.DataFrame) -> list[dict]:
    site = _identity_values(frame, "店铺-站点")
    sku = _identity_values(frame, "积加SKU")
    msku = _identity_values(frame, "MSKU")
    empty_site = site.eq("")
    empty_sku = sku.eq("")

    keys = pd.DataFrame({"site": site, "sku": sku, "msku": msku})
    valid_group = site.ne("") & sku.ne("")
    multiple_rows = valid_group & keys.duplicated(["site", "sku"], keep=False)
    duplicate_full_key = (
        valid_group
        & msku.ne("")
        & keys.duplicated(["site", "sku", "msku"], keep=False)
    )
    duplicate_msku = duplicate_full_key | (multiple_rows & msku.eq(""))

    scale = _text_values(frame, "规模定位")
    stocking = _text_values(frame, "备货定位")
    days = _text_values(frame, "已下单可售天数")
    unknown_scale = ~scale.isin(_KNOWN_SCALES)
    non_numeric_days = pd.to_numeric(days, errors="coerce").isna()

    issues: list[dict] = []
    _append_issue(
        issues,
        severity="error",
        code="empty_site",
        message="店铺-站点不能为空",
        mask=empty_site,
    )
    _append_issue(
        issues,
        severity="error",
        code="empty_sku",
        message="积加SKU不能为空",
        mask=empty_sku,
    )
    _append_issue(
        issues,
        severity="error",
        code="duplicate_msku",
        message="同一店铺-站点和积加SKU下的 MSKU 必须非空且唯一",
        mask=duplicate_msku,
    )
    _append_issue(
        issues,
        severity="warning",
        code="unknown_scale",
        message="规模定位必须为短尾、中尾或长尾",
        mask=unknown_scale,
    )
    _append_issue(
        issues,
        severity="warning",
        code="empty_stocking",
        message="备货定位不能为空",
        mask=stocking.eq(""),
    )
    _append_issue(
        issues,
        severity="warning",
        code="non_numeric_days",
        message="已下单可售天数必须为数值",
        mask=non_numeric_days,
    )
    return issues


def _position_records(
    frame: pd.DataFrame,
) -> dict[tuple[str, str, str], list[tuple[Any, ...]]]:
    records: dict[tuple[str, str, str], list[tuple[Any, ...]]] = {}
    normalized_keys = pd.DataFrame(
        {column: _identity_values(frame, column) for column in POSITION_KEY}
    )
    for key_values, source_values in zip(
        normalized_keys.itertuples(index=False, name=None),
        frame[_POSITION_VALUES].itertuples(index=False, name=None),
    ):
        key = tuple(key_values)
        values = tuple(_position_comparison_text(value) for value in source_values)
        records.setdefault(key, []).append(values)
    return records


def _position_comparison_text(value: Any) -> str:
    safe_value = _json_safe(value)
    if safe_value is None:
        return ""
    if isinstance(safe_value, float) and safe_value.is_integer():
        return str(int(safe_value))
    return str(safe_value)


def position_diff(base: pd.DataFrame, candidate: pd.DataFrame) -> dict[str, int]:
    base_records = _position_records(base)
    candidate_records = _position_records(candidate)
    added = 0
    modified = 0
    deleted = 0
    unchanged = 0
    for key in base_records.keys() | candidate_records.keys():
        base_remaining = list(base_records.get(key, []))
        candidate_unmatched = []
        for candidate_values in candidate_records.get(key, []):
            try:
                match_index = base_remaining.index(candidate_values)
            except ValueError:
                candidate_unmatched.append(candidate_values)
            else:
                unchanged += 1
                base_remaining.pop(match_index)
        paired = min(len(base_remaining), len(candidate_unmatched))
        modified += paired
        deleted += len(base_remaining) - paired
        added += len(candidate_unmatched) - paired
    return {
        "added": added,
        "modified": modified,
        "deleted": deleted,
        "unchanged": unchanged,
    }


def position_change_warnings(
    reference: pd.DataFrame,
    candidate: pd.DataFrame,
) -> list[dict]:
    metrics = (
        ("row_count", "行数", len(reference), len(candidate)),
        (
            "sites",
            "站点数",
            int(_identity_values(reference, "店铺-站点").replace("", pd.NA).nunique()),
            int(_identity_values(candidate, "店铺-站点").replace("", pd.NA).nunique()),
        ),
        (
            "skus",
            "积加 SKU 数",
            int(_identity_values(reference, "积加SKU").replace("", pd.NA).nunique()),
            int(_identity_values(candidate, "积加SKU").replace("", pd.NA).nunique()),
        ),
    )
    warnings = []
    for code, label, before, after in metrics:
        if before == after:
            continue
        if before > 0 and after == 0:
            issue_code = f"{code}_cleared"
            message = f"{label}从 {before} 清空为 0"
        elif before == 0 or abs(after - before) / before >= 0.5:
            issue_code = f"{code}_changed"
            message = f"{label}从 {before} 变为 {after}，变化达到或超过 50%"
        else:
            continue
        warnings.append(
            {
                "severity": "warning",
                "code": issue_code,
                "message": message,
                "row_numbers": [],
                "before": before,
                "after": after,
            }
        )
    return warnings


def write_position_workbook(path: Path, frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MSKU_视图"
    sheet.append(POSITION_SOURCE_COLUMNS)
    for values in frame[POSITION_SOURCE_COLUMNS].itertuples(index=False, name=None):
        sheet.append([None if pd.isna(value) else value for value in values])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
