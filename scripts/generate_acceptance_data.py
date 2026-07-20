from __future__ import annotations

import argparse
from pathlib import Path
import sys

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from delivery_note.pipeline import IMPORT_COLUMNS, POSITION_SOURCE_COLUMNS


def _write_table(path: Path, headers: list[str], row: list[object]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(row)
    workbook.save(path)


def _write_delivery(path: Path, quantity: int) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "汇总"
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(["SKU", "US站", "总计"])
    sheet.append(["SKU-A", quantity, quantity])
    workbook.save(path)


def _write_position(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MSKU_视图"
    sheet.append(POSITION_SOURCE_COLUMNS)
    sheet.append(["SEEKWAY:US", "SKU-A", "MSKU-A", "短尾", "备货", 90])
    workbook.save(path)


def _write_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "交货导入模板"
    sheet.merge_cells("A1:G1")
    sheet.append(IMPORT_COLUMNS)
    sheet.append(["示例仓", "GYS-023", "SKU-A", 1, "SEEKWAY:US", "", ""])
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    workbook.save(path)


def generate_acceptance_data(output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "purchase": output_dir / "purchase.xlsx",
        "product": output_dir / "product.xlsx",
        "supplier": output_dir / "supplier.xlsx",
        "position": output_dir / "position.xlsx",
        "template": output_dir / "template.xlsx",
        "delivery_first": output_dir
        / "260720-KuangBiao-A交货单-发货1箱.xlsx",
        "delivery_second": output_dir
        / "260720-KuangBiao-B交货单-发货1箱.xlsx",
    }

    _write_table(
        paths["purchase"],
        ["单据状态", "供应商", "SKU", "平台站点", "目的仓", "未交量"],
        [
            "待交货",
            "KuangBiao",
            "SKU-A",
            "AMAZON:SEEKWAY:US",
            "供应商成品本地仓",
            100,
        ],
    )
    _write_table(
        paths["product"],
        ["SKU", "店铺/站点", "品类A", "锁仓MKSU"],
        ["SKU-A", "SEEKWAY:US", "水鞋", "锁"],
    )
    _write_table(
        paths["supplier"],
        ["供应商编号", "供应商名称", "状态"],
        ["GYS-023", "KuangBiao", "启用"],
    )
    _write_position(paths["position"])
    _write_template(paths["template"])
    _write_delivery(paths["delivery_first"], 80)
    _write_delivery(paths["delivery_second"], 80)
    return paths


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="生成脱敏的 Web 验收 Excel 数据")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("acceptance_data"),
        help="输出目录",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    paths = generate_acceptance_data(args.output_dir)
    for kind, path in paths.items():
        print(f"{kind}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
